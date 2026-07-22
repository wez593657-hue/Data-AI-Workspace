"""Synchronize only the two deadline ADS Mapping targets from procedure evidence."""

from __future__ import annotations

import hashlib
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT / "data_assets/mapping/dws_to_ads/ADS应用层数据模型_CRM_ V1.0.xlsx"
MARKDOWN = ROOT / "data_assets/mapping/dws_to_ads/dws到ads映射.md"


def item(table: str, alias: str, field: str, name: str, rule: str) -> tuple[str, str, str, str, str]:
    return table, alias, field, name, rule


DETAIL = {
    "PERSN_LEGAL_BK_CODE": item("DWD_CUST_INDV_INFO", "ci", "PERSN_LEGAL_BK_CODE", "法人行号", "直接取客户基本信息法人行号"),
    "DATA_DATE": item("TMP_CDR_DTL_DUE_WIN", "w", "END_DT", "统计周期结束日期", "按 M/Q/N 周期生成，格式化为 yyyymmdd"),
    "CUST_ID": item("TMP_CDR_DTL_DUE_WIN", "w", "CUST_ID", "客户编号", "直接取到期窗口客户编号"),
    "CUST_NAME": item("TMP_CDR_DTL_CUST_BASE", "cb", "CUST_NAME", "客户名称", "直接取客户基础中间表"),
    "CUST_LVL": item("TMP_CDR_DTL_CUST_BASE", "cb", "CUST_LVL", "客户等级", "由 DWD_CUST_INDV_INFO.CUST_HRAKY 映射为 CUST_LVL 后取值"),
    "DEPO_CURNT_DEPO_BAL": item("TMP_CDR_DTL_CUST_BASE", "cb", "DEPO_CURNT_DEPO_BAL", "活期余额", "直接取客户余额中间表"),
    "FIXD_DEPO_BAL": item("TMP_CDR_DTL_CUST_BASE", "cb", "FIXD_DEPO_BAL", "定期余额", "直接取客户余额中间表"),
    "FIN_AMT": item("TMP_CDR_DTL_CUST_BASE", "cb", "FIN_AMT", "理财余额", "直接取客户余额中间表"),
    "STAT_PERD": item("TMP_CDR_DTL_DUE_WIN", "w", "STAT_PERD", "统计周期", "直接取到期窗口统计周期"),
    "STATIS_TYP": item("TMP_CDR_DTL_DUE_WIN", "w", "STATIS_TYP", "承接类型", "1 存款、2 理财、0 汇总"),
    "EXPR_AMT": item("TMP_CDR_DTL_DUE_WIN", "w", "EXPR_AMT", "已到期金额", "直接取窗口已到期金额"),
    "MATURE_TTL_AMT": item("TMP_CDR_DTL_DUE_WIN", "w", "MATURE_TTL_AMT", "到期总金额", "直接取窗口到期总金额"),
    "TAKE_RATE": item("TMP_CDR_DTL_TAKE_AMT", "t", "TAKE_AMT_30D", "30天承接金额", "ROUND(NVL(t.TAKE_AMT_30D,0)/w.EXPR_AMT*100,2)，EXPR_AMT=0 时为 0"),
    "FIX_DEPO_MATURE_AMT": item("TMP_CDR_DTL_DUE_WIN", "w", "EXPR_AMT", "到期金额", "STATIS_TYP=1 时取 EXPR_AMT，否则为 0"),
    "FIX_DEPO_MATURE_TTL_AMT": item("TMP_CDR_DTL_DUE_WIN", "w", "MATURE_TTL_AMT", "到期总金额", "STATIS_TYP=1 时取 MATURE_TTL_AMT，否则为 0"),
    "FIX_DEPO_TAKE_RATE": item("TMP_CDR_DTL_TAKE_AMT", "t", "BUY_DEPO_AMT_30D", "30天购买定期存款金额", "STATIS_TYP=1 时按 BUY_DEPO_AMT_30D/EXPR_AMT 计算"),
    "CNTCT_STATE": item("ADS_MKT_REC_INFO", "m", "CUST_ID,MKT_TIME", "客户营销记录", "存在 MKT_TIME 且不晚于跑批日则为 1，否则为 0"),
    "UNDTAKE_STATE": item("TMP_CDR_DTL_TAKE_AMT", "t", "TAKE_AMT_30D", "30天承接金额", "TAKE_AMT_30D/EXPR_AMT >= 0.8 则为 1，否则为 0"),
    "FIXED_FIN_MATURE_TRAN_INSUR_AMT": item("TMP_CDR_DTL_TAKE_AMT", "t", "BUY_INSUR_AMT_30D", "30天购买保险金额", "仅作为保险转化金额；保险不计入 TAKE_AMT_30D"),
    "FIN_MATURE_TRAN_FIXED_AMT": item("TMP_CDR_DTL_TAKE_AMT", "x", "BUY_DEPO_AMT_30D", "30天购买定期存款金额", "客户维度汇总 x.STATIS_TYP=2 的 BUY_DEPO_AMT_30D"),
    "FIXED_MATURE_TRAN_FIN_AMT": item("TMP_CDR_DTL_TAKE_AMT", "x", "BUY_FIN_AMT_30D", "30天购买理财金额", "客户维度汇总 x.STATIS_TYP=1 的 BUY_FIN_AMT_30D"),
    "FRST_MATURE_PK_BF_DAY_AUM_BAL": item("TMP_CDR_DTL_AUM_BAL", "ap", "AUM_BAL", "到期前一日AUM余额", "取 AUM_TYP=PREV 的 AUM_BAL"),
    "LAST_END_DATE": item("TMP_CDR_DTL_DUE_WIN", "w", "LAST_EXPR_DT", "最后一笔到期日期", "格式化为 yyyymmdd"),
    "POST_ID": item("TMP_CDR_DTL_CUST_BASE", "cb", "POST_ID", "管户经理职位编号", "直接取客户基础中间表"),
    "ORG_ID": item("TMP_CDR_DTL_CUST_BASE", "cb", "ORG_ID", "归属机构", "直接取客户基础中间表"),
}

STATIS = {
    "PERSN_LEGAL_BK_CODE": item("TMP_CDR_STAT_SRC", "s", "PERSN_LEGAL_BK_CODE", "法人行号", "直接取统计来源中间表"),
    "DATA_DATE": item("TMP_CDR_STAT_SRC", "s", "DATA_DATE", "数据日期", "直接取统计来源中间表"),
    "STATIS_OBJ": item("TMP_CDR_STAT_SRC", "s", "STATIS_OBJ", "统计对象", "机构维度取机构层级展开结果，客户经理维度取 POST_ID"),
    "STATIS_CYCLE": item("TMP_CDR_STAT_SRC", "s", "STAT_PERD", "统计周期", "由明细统计周期映射为统计周期"),
    "STATIS_TYP": item("TMP_CDR_STAT_SRC", "s", "STATIS_TYP", "承接类型", "直接取统计来源中间表"),
    "EXPR_CUST_CNT": item("TMP_CDR_STAT_SRC", "s", "EXPR_AMT,CUST_ID", "已到期金额、客户编号", "COUNT(DISTINCT CUST_ID) WHERE EXPR_AMT>0"),
    "TTL_EXPR_CUST_CNT": item("TMP_CDR_STAT_SRC", "s", "MATURE_TTL_AMT,CUST_ID", "到期总金额、客户编号", "COUNT(DISTINCT CUST_ID) WHERE MATURE_TTL_AMT>0"),
    "EXPR_AMT": item("TMP_CDR_STAT_SRC", "s", "EXPR_AMT", "已到期金额", "SUM(EXPR_AMT)"),
    "TTL_EXPR_AMT": item("TMP_CDR_STAT_SRC", "s", "MATURE_TTL_AMT", "到期总金额", "SUM(MATURE_TTL_AMT)"),
    "CUST_UNDTAKE_RATE": item("TMP_CDR_STAT_SRC", "s", "CUST_TAKE_FLG,CUST_ID,EXPR_AMT", "客户承接标志、客户编号、已到期金额", "承接客户数/已到期客户数*100"),
    "ASSET_KEEP_RATE": item("TMP_CDR_STAT_SRC", "s", "CURR_AUM_BAL,FRST_MATURE_PK_BF_DAY_AUM_BAL", "当前AUM、到期前一日AUM", "SUM(CURR_AUM_BAL)/SUM(FRST_MATURE_PK_BF_DAY_AUM_BAL)*100"),
    "ASSET_UNDTAKE_RATE": item("TMP_CDR_STAT_SRC", "s", "EXPR_AMT,TAKE_RATE_30D", "已到期金额、30天承接率", "SUM(EXPR_AMT*TAKE_RATE_30D/100)/SUM(EXPR_AMT)*100"),
    "DEPO_TO_FIN_CONVRS_RATE": item("TMP_CDR_STAT_SRC", "s", "FIXED_MATURE_TRAN_FIN_AMT,EXPR_AMT", "定期到期转理财金额、到期金额", "SUM(FIXED_MATURE_TRAN_FIN_AMT)/SUM(EXPR_AMT)*100"),
    "INSUR_CONVRS_RATE": item("TMP_CDR_STAT_SRC", "s", "FIXED_FIN_MATURE_TRAN_INSUR_AMT,EXPR_AMT", "理财到期转保险金额、到期金额", "SUM(FIXED_FIN_MATURE_TRAN_INSUR_AMT)/SUM(EXPR_AMT)*100"),
    "FIN_TO_DEPO_CONVRS_RATE": item("TMP_CDR_STAT_SRC", "s", "FIN_MATURE_TRAN_FIXED_AMT,EXPR_AMT", "理财到期转定期金额、到期金额", "SUM(FIN_MATURE_TRAN_FIXED_AMT)/SUM(EXPR_AMT)*100"),
}


def clean(value: object) -> str:
    return "" if value is None else str(value).strip().replace("|", "&#124;").replace("\n", "<br>")


def fill_excel() -> dict[str, int]:
    workbook = load_workbook(WORKBOOK)
    counts: dict[str, int] = {}
    for worksheet in workbook.worksheets:
        table = clean(worksheet.cell(2, 6).value)
        values = DETAIL if table == "ADS_CUST_DEADLINE_RMND_DTL" else STATIS if table == "ADS_CUST_DEADLINE_RMND_STATIS" else None
        if values is None:
            continue
        count = 0
        for row in range(6, worksheet.max_row + 1):
            field = clean(worksheet.cell(row, 2).value)
            if field == "变更登记":
                break
            if field not in values:
                continue
            source_table, alias, source_field, source_name, rule = values[field]
            for column, value in ((12, source_table), (13, alias), (14, source_field), (15, source_name), (16, rule)):
                worksheet.cell(row, column).value = value
            count += 1
        counts[table] = counts.get(table, 0) + count
    workbook.save(WORKBOOK)
    return counts


def render_section(table: str, values: dict[str, tuple[str, str, str, str, str]], workbook) -> str:
    worksheet = next(ws for ws in workbook.worksheets if clean(ws.cell(2, 6).value) == table)
    lines = [f"### {table}", "", "| 目标字段 | 目标字段中文名 | 目标字段类型 | 源系统表名 | 源表别名 | 源系统字段英文名 | 源系统字段中文名 | 映射规则 |", "|----------|----------------|--------------|------------|----------|------------------|------------------|----------|"]
    for row in range(6, worksheet.max_row + 1):
        field = clean(worksheet.cell(row, 2).value)
        if field == "变更登记":
            break
        if field in values:
            lines.append("| " + " | ".join(clean(worksheet.cell(row, col).value) for col in (2, 5, 4, 12, 13, 14, 15, 16)) + " |")
    return "\n".join(lines)


def update_markdown() -> None:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=False)
    text = MARKDOWN.read_text(encoding="utf-8")
    digest = hashlib.sha256(WORKBOOK.read_bytes()).hexdigest()
    text = re.sub(r"(?m)^- Excel SHA-256:.*$", f"- Excel SHA-256: `{digest}`", text)
    for table, values in (("ADS_CUST_DEADLINE_RMND_DTL", DETAIL), ("ADS_CUST_DEADLINE_RMND_STATIS", STATIS)):
        replacement = render_section(table, values, workbook)
        pattern = rf"(?ms)^### {table}\n.*?(?=^### |^---$|\Z)"
        text, count = re.subn(pattern, replacement + "\n\n", text, count=1)
        if count != 1:
            raise RuntimeError(f"未找到唯一目标章节: {table}")
    MARKDOWN.write_text(text, encoding="utf-8", newline="\n")


if __name__ == "__main__":
    print(fill_excel())
    update_markdown()
