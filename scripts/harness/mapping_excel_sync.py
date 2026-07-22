"""Synchronize the three Mapping Markdown files from their authoritative Excel models."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from .evidence_store import sha256_file


WORKBOOKS = {
    "ods_to_dwd": "DWD明细层数据模型_CRM_ V1.0.xlsx",
    "dwd_to_dws": "DWS汇总层数据模型_CRM_ V1.0.xlsx",
    "dws_to_ads": "ADS应用层数据模型_CRM_ V1.0.xlsx",
}
OUTPUTS = {
    "ods_to_dwd": "ods到dwd映射.md",
    "dwd_to_dws": "dwd到dws映射.md",
    "dws_to_ads": "dws到ads映射.md",
}


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("|", "&#124;").replace("\r\n", "<br>").replace("\n", "<br>")


def _field_rows(ws: Any) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    started = False
    for row in ws.iter_rows(min_row=6, values_only=True):
        target = _clean(row[1] if len(row) > 1 else None)
        candidate_c = _clean(row[2] if len(row) > 2 else None)
        candidate_d = _clean(row[3] if len(row) > 3 else None)
        data_type = candidate_c if re.match(r"^(VARCHAR2?|NUMBER|NUMERIC|DATE|TIMESTAMP|CHAR|BPCHAR)", candidate_c, re.IGNORECASE) else candidate_d
        if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_$]*", target) or not data_type:
            if started:
                break
            continue
        started = True
        records.append({
            "target_field": target,
            "target_type": data_type,
            "target_name": _clean(row[4] if len(row) > 4 else None),
            "source_table": _clean(row[11] if len(row) > 11 else None),
            "source_field": _clean(row[13] if len(row) > 13 else None),
            "rule": _clean(row[15] if len(row) > 15 else None),
        })
    return records


def extract_workbook(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=False, data_only=False)
    records: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        target_table = _clean(worksheet.cell(2, 6).value)
        if not target_table:
            continue
        for row in _field_rows(worksheet):
            records.append({"target_table": target_table, **row, "sheet": worksheet.title, "workbook": str(path), "workbook_sha256": sha256_file(path)})
    return records


def _render(flow: str, workbook: Path, records: list[dict[str, Any]]) -> str:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for record in records:
        grouped.setdefault(record["target_table"].strip().lower(), []).append(record)
    lines = [f"# {flow.replace('_to_', '到').upper()}映射 字段映射", "", "## 映射来源", "", f"- Excel：`{workbook.as_posix()}`", f"- Excel SHA-256：`{sha256_file(workbook)}`", "", "## 映射概览", "", "| 目标表 | 字段数 |", "|--------|-------:|"]
    lines.extend(f"| {table.upper()} | {len(rows)} |" for table, rows in sorted(grouped.items()))
    lines.extend(["", "## 字段映射详情", ""])
    for table, rows in grouped.items():
        lines.extend([f"### {rows[0]['target_table']}", "", "| 目标字段 | 目标字段中文名 | 目标字段类型 | 源表 | 源字段 | 映射规则 |", "|----------|----------------|--------------|------|--------|----------|"])
        for row in rows:
            lines.append(f"| {row['target_field']} | {row['target_name']} | {row['target_type']} | {row['source_table']} | {row['source_field']} | {row['rule']} |")
        lines.append("")
    lines.extend(["---", "", "*本文件由对应 Excel 模型同步生成；Excel 更新后必须重新生成本文件。", ""])
    return "\n".join(lines)


def sync_mapping_markdown(root: Path) -> list[dict[str, Any]]:
    outputs: list[dict[str, Any]] = []
    for flow, workbook_name in WORKBOOKS.items():
        directory = root / "data_assets" / "mapping" / flow
        workbook = directory / workbook_name
        output = directory / OUTPUTS[flow]
        if not workbook.exists():
            raise FileNotFoundError(workbook)
        records = extract_workbook(workbook)
        output.write_text(_render(flow, workbook.relative_to(root), records), encoding="utf-8", newline="\n")
        outputs.append({"flow": flow, "workbook": str(workbook.relative_to(root)), "output": str(output.relative_to(root)), "record_count": len(records), "sha256": sha256_file(workbook)})
    return outputs
