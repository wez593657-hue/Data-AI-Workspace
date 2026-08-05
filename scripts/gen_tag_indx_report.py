import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

hdr_font = Font(name='Microsoft YaHei', bold=True, size=10, color='FFFFFF')
hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
cat_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
red_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
title_font = Font(name='Microsoft YaHei', bold=True, size=14, color='1F4E79')
sub_font = Font(name='Microsoft YaHei', bold=True, size=11, color='1F4E79')
normal_font = Font(name='Microsoft YaHei', size=10)
bold_font = Font(name='Microsoft YaHei', bold=True, size=10)
green_font = Font(name='Microsoft YaHei', size=10, color='006600')
red_font = Font(name='Microsoft YaHei', bold=True, size=10, color='CC0000')
orange_font = Font(name='Microsoft YaHei', size=10, color='BF8F00')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
wrap_top = Alignment(wrap_text=True, vertical='top')
wrap_center = Alignment(wrap_text=True, vertical='top', horizontal='center')

def style_hdr(ws, r, cols):
    for c in range(1, cols+1):
        cl = ws.cell(row=r, column=c)
        cl.font = hdr_font; cl.fill = hdr_fill; cl.alignment = wrap_center; cl.border = thin_border

def style_cell(ws, r, c, val, font=None, fill=None, align=None):
    cl = ws.cell(row=r, column=c, value=val)
    cl.font = font or normal_font
    cl.alignment = align or wrap_top
    cl.border = thin_border
    if fill: cl.fill = fill

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def section_row(ws, r, cols, text, fill=cat_fill):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    style_cell(ws, r, 1, text, sub_font, fill)
    ws.row_dimensions[r].height = 24

# ===== Sheet 1: 标签-指标映射矩阵 =====
ws1 = wb.active
ws1.title = '标签指标映射矩阵'

ws1.merge_cells('A1:H1')
style_cell(ws1, 1, 1, '标签 → 指标映射矩阵', Font(name='Microsoft YaHei', bold=True, size=14, color='1F4E79'))
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 28

ws1.merge_cells('A2:H2')
style_cell(ws1, 2, 1, '版本: v1.0 | 日期: 2026-08-04 | 源文件: 指标.xlsx · 标签Sheet (125行)', Font(name='Microsoft YaHei', size=9, color='666666'))
ws1['A2'].alignment = Alignment(horizontal='center')

h1 = ['标签行号', '标签名称', '标签类目', '标签口径(摘要)', '映射指标', '匹配度', '取数可行性', '说明']
for c, h in enumerate(h1, 1):
    style_cell(ws1, 4, c, h, hdr_font, hdr_fill, wrap_center)
ws1.row_dimensions[4].height = 22

mappings = [
    # === DIRECT mappings ===
    ['R21', '是否睡眠户', '70_客户风险', 'AUM<100 AND 一年内无主动动账', '—(睡眠户唤醒模块)', '高', '直接复用', '与睡眠户唤醒明细模块的睡眠判定规则完全一致，AUM+动账双条件'],
    ['R22', 'CRM客户层级', '20_客户价值', '上月AUM月日均计算的客户等级(0-10级)', '0052/0053/0054', '高', '直接复用', '与DWS_CUST_LVL_INFO口径一致，可直接替代或交叉校验'],
    ['R23', '客户等级', '20_客户价值', '6级分类(大众/潜客/优质/财富/贵宾/私行)', '0052/0053/0054', '中', '间接辅助', '6级分类与DWS的7级映射不同，需确认映射关系后才能做等级增量计算'],
    ['R24', 'AUM余额', '20_客户价值', '理财+存款+保险等余额汇总', '0052算', '高', '直接复用', '对应DWS_CUST_ASSE_LIAB.AUM_BAL，可做交叉校验'],
    ['R25', 'AUM月日均', '20_客户价值', '当前(当月)AUM月日均', '0050/0051/0063', '高', '直接复用', '对应DWS_CUST_ASSE_LIAB BAL_TYPE=2，临界客户(0063)当月AUM月日均可直接用'],
    ['R27', '存款余额', '20_客户价值', '个人存款产品余额汇总', '0046-0049', '高', '直接复用', '对应储蓄存款余额(需确认是否含乐惠存+大额存单)'],
    ['R28', '存款月日均余额', '20_客户价值', '当前月存款日均', '0051', '高', '直接复用', '对应储蓄存款月日均(BAL_TYPE=2)'],
    ['R29', '活期存款余额', '10_产品持有', '活期存款+智能存款活期', '0046/0068', '高', '直接复用', '对应DEPO_CURNT_DEPO_BAL'],
    ['R32', '是否持有定期存款产品', '10_产品持有', '定期存款账户余额>0', '0046', '中', '间接辅助', '可做0052/0053/0054等级变化的归因分析'],
    ['R33', '定期存款余额', '10_产品持有', '普通定期+智能定期', '0046-0049', '高', '直接复用', '对应FIXD_DEPO_BAL'],
    ['R38-R41', '乐惠存产品(是否/余额/月日均)', '10_产品持有', '乐惠存余额', '0046', '高', '直接复用', '对应LEHUI_BAL，DWS_CUST_ASSE_LIAB中也有此字段'],
    ['R44-R47', '大额存单(是否/余额/月日均)', '10_产品持有', '大额存单余额', '0046', '高', '直接复用', '对应LARGEDP_BAL'],
    ['R50/R53/R54', '持有理财产品/余额/月日均', '10_产品持有', '理财余额(自营+代销)', '0055-0059', '高', '直接复用', '对应DWS_CUST_ASSE_LIAB.FIN_BAL'],
    ['R63-R66', '代销理财(是否/余额/月日均)', '10_产品持有', '代销理财产品', '0058-0060', '高', '直接复用', '代销大类筛选后可对应0058-0060'],
    ['R67', '是否持有保险产品', '10_产品持有', '保费>0且未退保', '0061', '高', '直接复用', '对应DWD_ACCT_INSUR,与0061口径高度一致'],
    ['R68', '保险产品首期保费', '10_产品持有', '首期保费', '0061', '高', '直接复用', '对应NEW_INSUR_AMT, 0061核心计算直接使用'],
    ['R71', '贷款余额', '10_产品持有', '客户所有借据余额之和', '0062', '高', '直接复用', '对应LOAN_BAL, 0062核心字段'],
    ['R69', '是否贷款客户', '10_产品持有', '授信额度>0且合同有效期>=当前', '—', '中', '间接辅助', '可用于0062的客户筛选范围'],
    ['R84', '近1月累计交易笔数', '40_交易行为', '主动动账笔数', '—', '中', '间接辅助', '对应DWD_TX_ASET交易数据，可用于交叉校验指标趋势'],
    ['R85', '近1月累计交易金额', '40_交易行为', '主动动账金额', '—', '中', '间接辅助', '同上'],
    ['R89', '近1月第三方支付转出笔数', '40_交易行为', '网联渠道', '—', '中', '间接辅助', '可对应标签表网联渠道交易统计'],
    ['R90', '近1月第三方支付转出金额', '40_交易行为', '网联渠道', '—', '中', '间接辅助', '同上'],

    # === Q1-related (merchant) ===
    ['R12', '是否一码付商户', '10_产品持有', '办理一码付业务', '0068/0069', '高', '直接复用', '一码付=条码收单，0068/0069核心客群筛选'],
    ['R120', '是否收单商户小微商户', '00_基础信息', '一码付商户-小微商户', '0068/0069', '高', '**Q1直接答案**', 'mct_type=smallBusinesses → 小微商户, 可直接用于0068/0069的商户类型区分！'],
    ['R121', '是否收单商户个体商户', '00_基础信息', '一码付商户-个体商户', '0068/0069', '高', '**Q1直接答案**', 'mct_type=personage → 个体商户, 可直接用于0068/0069的商户类型区分！'],
    ['R122', '收单商户上月交易笔数', '40_交易行为', '收单商户上月交易笔数', '0068', '高', '直接复用', '部分对应0068近30天交易笔数条件(上月≠近30天,需确认)'],
    ['R123', '收单商户上月交易金额', '40_交易行为', '收单商户上月交易金额', '0068', '高', '直接复用', '部分对应0068近30天交易金额条件(同上)'],
    ['R124', '是否收单价值商户', '20_客户价值', '资产+交易量双达标', '0068', '高', '直接复用', '与0068价值商户定义完全一致! AUM/活期+条码收单双条件'],

    # === Q2-related (card binding) ===
    ['R107', '是否登录手机银行并完成绑卡客户', '50_渠道分析', '借记卡绑卡', '0070/0071', '中', '间接辅助', '绑卡行为标签, 但不区分三方平台(微信/支付宝/云闪付)'],
    ['R110', '是否完成三方支付绑卡', '50_渠道分析', '三方支付绑卡+渠道筛选', '0070/0071', '高', '**Q2半答案**', '明确支持渠道筛选(微信/支付宝/云闪付/美团/京东), 但"首次"绑定的具体判断逻辑仍需确认(首次=第一笔?). 时间范围可自定义'],
    ['R117', '是否为手机银行月活跃客户', '50_渠道分析', '当月动账>=6笔', '—', '中', '间接辅助', '手机银行活跃定义, 与0067口径不同但可参考'],
    ['R118', '是否为微信绑卡交易月活跃客户', '50_渠道分析', '当月微信动账>=6笔', '—', '中', '间接辅助', '微信绑卡活跃, 可辅助0070绑卡指标'],

    # === Q3-related (mobile banking active) ===
    ['R109', '当月登录手机银行次数', '50_渠道分析', '当月登录次数', '0067', '高', '直接复用', '可直接用于0067的活跃统计, 活跃阈值(>=1次?)待确认'],
    ['R116', '连续登录手机银行天数', '50_渠道分析', '连续登录天数', '0067', '中', '间接辅助', '更高门槛的活跃定义'],
    ['R108', '注册并登录手机银行客户的天数', '50_渠道分析', '首次注册起算天数', '0067', '中', '间接辅助', '客户手机银行生命周期'], 

    # === Q4-related (payroll) ===
    ['R18', '是否代发客户', '00_基础信息', '代发工资/惠民惠农/社保/国库', '0064', '高', '直接复用', '涵盖代发薪全部子类型, 0064可直接用此标签判定'],
    ['R72', '是否代发工资客户', '20_客户价值', '上月存在代发工资流水', '0064', '高', '直接复用', '0064新签代发客户≥上月有代发。代发工资子类与R18口径交叉验证'],
    ['R73', '上月代发工资金额', '20_客户价值', '上月代发金额', '—(辅助)', '中', '间接辅助', '可与0064交叉验证'],
    ['R74', '最近一次代发工资日期', '20_客户价值', '最近代发日期', '0064', '高', '直接复用', '可判断"新签"时间点(>=考核期初)'],
    ['R76', '代发工资累计金额_6', '20_客户价值', '近6月累计', '—(辅助)', '中', '间接辅助', '可用于R14近半年月均收入计算'],

    # === Indicator-directed matches ===
    ['R26', 'AUM季日均', '20_客户价值', '当前季AUM季日均', '—(辅助)', '中', '间接辅助', '可用于0049较上季的口径验证'],
    ['R42', '最近一笔定期存款到期天数', '10_产品持有', '到期天数', '—(到期承接)', '高', '直接复用', '到期承接模块直接使用此标签或可替代ADM计算'],
    ['R43', '最近一笔定期存款到期金额', '10_产品持有', '到期金额', '—(到期承接)', '高', '直接复用', '同上'],
    ['R51', '最近一笔理财到期天数', '10_产品持有', '理财到期天数', '—(到期承接)', '高', '直接复用', '同上'],
    ['R52', '最近一笔理财到期金额', '10_产品持有', '理财到期金额', '—(到期承接)', '高', '直接复用', '同上'],
    ['R93', '储蓄产品期限偏好', '30_客户偏好', '各期限产品占比', '—', '低', '辅助洞察', '客户洞察用, 非指标直接相关'],
    ['R94', '行内贷款是否逾期', '70_客户风险', '任一借据逾期天数>0', '0066', '中', '间接辅助', '逾期与不良(五级分类)是两个概念, 不直接对应0066'],
    ['R95', '近24月是否存在贷款逾期', '70_客户风险', '罚息>0', '0066', '中', '间接辅助', '逾期历史≠不良形成率, 辅助风险评估'],
    ['R102', '是否当年首次等级升级客户', '20_客户价值', '以上年12月31日为基数', '0052-0054', '中', '间接辅助', '与0052-0054的"期初→期末"口径平行, 可交叉验证'],
    ['R103', '首次升级客户升级等级', '20_客户价值', '升级到的等级', '0052-0054', '中', '间接辅助', '同上, 可验证等级增量口径'],
    ['R104', '是否当年等级降级客户', '20_客户价值', '年末→现在降级', '—', '低', '不直接相关', '0052-0054仅统计净增(不扣减降级)'],
    ['R106', '是否为高净值临界客户', '20_客户价值', '临界区间客户', '0063', '高', '**直接复用**', '与0063临界客户区间定义完全一致! [4.5,5万)等'],
    ['R125', '是否美团贷款客户', '—', '历史和现在持有', '0062', '高', '直接复用', '0062需排除美团贷款, 此标签可直接用于过滤'],
    ['R126', '是否微粒贷客户', '—', '历史和现在持有', '0062', '高', '直接复用', '0062需排除微粒贷, 此标签可直接用于过滤'],
]

row = 5
for item in mappings:
    for c, v in enumerate(item, 1):
        fl = None
        f = normal_font
        if item[5] == '高': fl = green_fill
        elif item[5] == '中': fl = yellow_fill
        style_cell(ws1, row, c, v, f, fl if fl else (None if row % 2 == 0 else PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')))
    # risk column highlight
    if 'Q1直接答案' in item[6] or 'Q2半答案' in item[6] or 'Q4' in item[0] or '排除' in item[6]:
        ws1.cell(row=row, column=7).font = red_font
    # match degree highlight
    if item[5] == '高':
        ws1.cell(row=row, column=6).font = green_font
        ws1.cell(row=row, column=6).fill = green_fill
    ws1.row_dimensions[row].height = 45
    row += 1

set_widths(ws1, [10, 34, 14, 45, 22, 10, 20, 55])

# ===== Sheet 2: Q问题答案发现 =====
ws2 = wb.create_sheet('Q问题关键发现')

ws2.merge_cells('A1:F1')
style_cell(ws2, 1, 1, '通过标签表发现的待确认问题关键线索', Font(name='Microsoft YaHei', bold=True, size=14, color='1F4E79'))
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 28

h2 = ['待确认问题', '原阻塞原因', '标签表找到的线索', '解决程度', '仍需确认', '建议行动']
for c, h in enumerate(h2, 1):
    style_cell(ws2, 3, c, h, hdr_font, hdr_fill, wrap_center)
ws2.row_dimensions[3].height = 22

q_findings = [
    ['Q1: 商户小微/个体区分',
     '不确定uepp_pay_mct_info.mct_type是否即为区分标准',
     'R120 是否收单商户小微商户: "通过商户类型区分小微商户"\nR121 是否收单商户个体商户: "通过商户类型区分个体商户"\nR124 是否收单价值商户: 明确区分小微(AUM>=2万/活期>=1万)和个体(AUM>=5万/活期>=2.5万)\n结论: mct_type=personage(个体) / smallBusinesses(小微) 是正确的区分标准',
     '**90%解决**',
     '1. 商户→CUST_ID映射方式(consumer_id? job_id?)\n2. AUM/活期指关联个人AUM还是商户结算存款?',
     '1. 可直接用R120/R121标签(如果标签表已入库)\n2. 或写存储过程时JOIN uepp_pay_mct_info, WHERE mct_type=对应值'],
    ['Q2: 绑卡"首次绑定"判断',
     'mbk_cust_acct无三方平台字段且无"首次"标志',
     'R110 是否完成三方支付绑卡: "可筛选绑卡的渠道(微信、支付宝、云闪付、美团支付、京东支付等), 筛选条件能否自定义时间范围"\n说明三方绑卡表支持渠道筛选!',
     '**60%解决**',
     '1. "首次"绑定的准确判定逻辑(该卡首次?该客户首次?)\n2. 渠道筛选的具体字段名\n3. mbk_cust_no→CUST_ID映射',
     '1. 找到标签表后台的绑卡源表(可能是mbk_cust_acct或独立绑卡表)\n2. 确认渠道字段名→渠道值映射\n3. 确认该表是否存储首次绑定时间'],
    ['Q3: 活跃频次阈值',
     '不确定手机银行登录"活跃"最低次数',
     'R109 当月登录手机银行次数: 当月值(自然月内登录次数)\nR116 连续登录手机银行天数\nR117 是否为手机银行月活跃客户: 当月动账>=6笔(含)\nR118 是否为微信绑卡交易月活跃客户: 当月微信动账>=6笔(含)',
     '**80%解决**',
     '1. 0067的"活跃"是否取登录>=1次?\n2. 与"月活跃"概念是否区分?',
     '可以参考: 0067按当年至少登录1次(=R109>=1 YTD), 月活跃按>=6笔(但月活跃是针对动账,0067是登录)'],
    ['Q4: 代发薪CTRAKT_TYP值域',
     '不确定CTRAKT_TYP字段代发以外的值',
     'R18 是否代发客户: 含行内代发工资、惠民惠农一卡通代发、省社保代发、市社保代发、国库集中支付、代发薪系统代发\nR72 是否代发工资客户: 上月代发工资流水(摘要为"代发工资")\n代发薪子类型已完整列出!',
     '**100%解决**',
     '需要确认DWD_CUST_SIGN_CTRAKT.CTRAKT_TYP字段是否有对应的这些值域编码',
     '直接用R18标签(包括所有子类型)替代直接查DWD_CUST_SIGN_CTRAKT\n或者按R18的子类型列表去确认CTRAKT_TYP的对应值'],
    ['Q5/Q6/Q8: 已解决',
     '之前已确认',
     '—',
     '100%',
     '—',
     '—'],
    ['Q7: 贵金属代销源表',
     '全量DDL搜索未发现贵金属表',
     '标签表中未发现"贵金属持有"/"贵金属代销收入"等标签\n说明贵金属代销在客户标签体系中也尚未覆盖',
     '0%',
     '仍需由贵金属业务负责人确认源表',
     '建议: 0065先做保险+理财, 贵金属代销待标签体系补齐后再补充'],
]

row = 4
for item in q_findings:
    for c, v in enumerate(item, 1):
        fl = None
        if '90%' in v or '100%' in v: fl = green_fill
        elif '60%' in v or '80%' in v: fl = yellow_fill
        elif '0%' in v: fl = red_fill
        style_cell(ws2, row, c, v, normal_font, fl if fl and c == 4 else None if c == 4 else (None if row % 2 == 0 else PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')))
    ws2.row_dimensions[row].height = 90
    row += 1

set_widths(ws2, [22, 32, 60, 14, 36, 50])

# ===== Sheet 3: 标签分类汇总 =====
ws3 = wb.create_sheet('标签分类汇总')

ws3.merge_cells('A1:E1')
style_cell(ws3, 1, 1, '标签表分类汇总 — 用于补充指标指标体系的数据维度', Font(name='Microsoft YaHei', bold=True, size=13, color='1F4E79'))
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 26

h3 = ['一级目录', '标签数量', '与指标相关标签', '新增分析维度', '取数建议']
for c, h in enumerate(h3, 1):
    style_cell(ws3, 3, c, h, hdr_font, hdr_fill, wrap_center)
ws3.row_dimensions[3].height = 22

cats = [
    ['00_基础信息类', '20个(R2-R21)', 'R18(代发客户), R120(小微商户), R121(个体商户)',
     '客户基本信息(年龄/性别/学历/职业/婚姻)可用于指标下的客户分层分析',
     'R18/R120/R121可直接用于0064/0068指标计算; 其他仅用于BI展示层的客户画像分析'],
    ['10_产品持有类', '30个(R12,R29-R71)', 'R12(一码付),R29(活期),R32/R33(定期),R38-R47(乐惠存/大额存单),R50-R66(理财/代销),R67/R68(保险),R71(贷款)',
     '各产品子类(开放式/封闭式理财,整存整取,乐惠存,大额存单)的余额和月日均, 可做产品归因分析',
     '多数标签与DWS/DWD源表字段直接对应, 用作交叉校验而非替代主数据源'],
    ['20_客户价值类', '15个(R14/R15/R22-R28/R72-R77/R98-R106/R124)', 'R22(客户层级),R24(AUM),R25(月日均),R27(存款),R72-R77(代发),R106(临界客户),R124(价值商户)',
     '客户等级升降级(R102-R105)、连续代发月数(R77)、最低权益等级(近3/6/9/12月, R98-R101)',
     'R102-R105可直接细化0052-0054的升级统计; R98-R101可做客户稳定性分析'],
    ['30_客户偏好类', '1个(R93)', '—',
     '储蓄产品期限偏好(R93): 各期限产品占储蓄总额比例',
     '仅用于BI分析/标签, 不直接对应指标体系'],
    ['40_客户交易行为特征', '10个(R84-R91,R111/R113-R115/R122-R123)', 'R84/R85(交易笔数/金额),R89/R90(三方支付),R122/R123(收单交易)',
     '第三方支付转出金额/笔数(R89/R90)、校园缴费(R111)、水电气缴费(R113/R114)、交易渠道偏好(R115)',
     '可用于指标数据趋势验证; 交易渠道偏好可用于产品推荐'],
    ['50_客户渠道分析类', '15个(R82-R83/R86-R88/R107-R119)', 'R86-R88(手机银行交易),R107(绑卡),R109(登录(0067)),R110(三方绑卡(Q2)),R117(月活跃),R118(微信活跃)',
     '手机银行消息阅读频率(R119)、热门活动次数(R112)、连续登录天数(R116)',
     'R109直接可用于0067(登录次数); R110直接可用于0070(Q2); R117/R118的"月活跃"定义可参考'],
    ['70_客户风险分析', '5个(R21/R91-R95)', 'R21(睡眠户),R94/R95(贷款逾期),R125(美团贷款),R126(微粒贷)',
     '资金过路户(R91:日均余额/累计转出<10%)、贷款逾期记录',
     'R125/R126直接可用于0062排除美团/微粒贷; R21与睡眠户唤醒模块口径完全一致'],
]

row = 4
for item in cats:
    for c, v in enumerate(item, 1):
        style_cell(ws3, row, c, v, normal_font, None if row % 2 == 0 else PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid'))
    ws3.row_dimensions[row].height = 75
    row += 1

set_widths(ws3, [22, 20, 55, 60, 55])

# ===== Sheet 4: 可从标签表直接取数的指标 =====
ws4 = wb.create_sheet('可从标签表取数的指标')

ws4.merge_cells('A1:G1')
style_cell(ws4, 1, 1, '可从标签表直接/间接取数的指标清单', Font(name='Microsoft YaHei', bold=True, size=13, color='1F4E79'))
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 26

h4 = ['指标编号', '指标名称', '取数方式', '对应标签', '数据提取路径', '数据一致性', '建议优先级']
for c, h in enumerate(h4, 1):
    style_cell(ws4, 3, c, h, hdr_font, hdr_fill, wrap_center)
ws4.row_dimensions[3].height = 22

direct_indicators = [
    # DIRECT - can fully use tag table
    ['0064', '代发薪客户净增', '直接取数', 'R18(是否代发客户), R72(是否代发工资客户), R74(最近一次代发日期)',
     '标签表 → 按CUST_ID+PERSN_LEGAL_BK_CODE JOIN → 判断考核期内新签(R74 BETWEEN 期初 AND 跑批日)',
     '高 — 标签口径覆盖代发薪全子类型', 'P0(可直接开发)'],
    ['0068', '支付结算年度价值商户数', '直接取数', 'R124(是否收单价值商户)',
     '标签表R124已预计算资产+交易量双达标; 如需拆分小微/个体, 加R120/R121区分',
     '高 — R124口径与0068完全一致', 'P0(可直接开发)'],
    ['0069', '二维码收款商户存款留存率', '直接取数', 'R120(小微), R121(个体), R122(交易笔数), R123(交易金额)',
     '标签表 → 筛选R120/R121 → 取R122/R123年累计交易金额 → 计算留存率',
     '中 — 年累计交易金额需确认(标签为"上月",0069为"当年")', 'P1(需确认时间窗口)'],
    ['0052/0053/0054', '财富/贵宾/私行客户数增量', '直接取数', 'R22(CRM客户层级), R102(当年首次升级), R103(升级等级)',
     '标签表 → R102=是 AND R103 IN (对应升级等级) → COUNT',
     '中 — R102以上年末为基数, 0052以活动/任务开始日为基数', 'P1(基期不同需修正)'],

    # INDIRECT - can use tag table for validation
    ['0062', '个人贷款净增额', '间接辅助', 'R125(美团贷款), R126(微粒贷), R71(贷款余额)',
     '排除: WHERE NOT R125 AND NOT R126 → 直接可排除美团/微粒贷客户\n余额: R71可做交叉验证',
     '高 — R125/R126直接解决美团/微粒贷排除问题!', 'P0(排除条件直接用标签)'],
    ['0067', '手机银行活跃客户数', '间接辅助', 'R109(当月登录次数), R117(月活跃,>=6笔)',
     'R109表天然统计登录次数 → 可按阈值(>=1次)汇总\n注意: R117的"月活跃"门槛更高(>=6笔动账)',
     '中 — 0067按"当年"计算, 标签按"当月", 需跨月聚合', 'P1(需确认0067活跃定义)'],
    ['0070/0071', '绑卡/绑卡率', '间接辅助', 'R110(完成三方支付绑卡), R107(手机银行绑卡)',
     'R110已标注支持渠道筛选+时间范围自定义 → 可直接用R110筛选\R110的渠道筛选字段可解决Q2中"哪个字段区分平台"',
     '中 — 需确认R110中"首次"的定义(卡级?客户级?)', 'P0(半解决Q2,确认后可直接用)'],
    ['0063', '临界客户净增', '直接取数', 'R106(是否为高净值临界客户)',
     '标签表R106已预计算: 临界区间与0063完全一致 → WHERE R106=是',
     '高 — R106口径与0063一致', 'P0(可直接开发)'],
]

row = 4
for item in direct_indicators:
    for c, v in enumerate(item, 1):
        fl = None
        if 'P0' in str(item[6]): fl = green_fill
        elif 'P1' in str(item[6]): fl = yellow_fill
        style_cell(ws4, row, c, v, normal_font, fl if fl and c == 7 else (None if row % 2 == 0 else PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')))
    ws4.row_dimensions[row].height = 80
    row += 1

set_widths(ws4, [14, 26, 14, 42, 55, 36, 24])

# ===== Sheet 5: 关键发现与行动建议 =====
ws5 = wb.create_sheet('关键发现与行动建议')

ws5.merge_cells('A1:B1')
style_cell(ws5, 1, 1, '标签表分析 - 关键发现与行动建议', Font(name='Microsoft YaHei', bold=True, size=14, color='1F4E79'))
ws5['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[1].height = 28

h5 = ['发现项', '详情与建议']
for c, h in enumerate(h5, 1):
    style_cell(ws5, 3, c, h, hdr_font, hdr_fill, wrap_center)
ws5.row_dimensions[3].height = 22

findings = [
    ('一、Q1(商户区分) 直接解决', 'R120(小微商户)和R121(个体商户)两个标签明确通过mct_type区分，R124(价值商户)按小微/个体分别设置了AUM/活期两个门槛。\n\n建议: Q1可以正式关闭。存储过程中使用 uepp_pay_mct_info.mct_type = "personage"/"smallBusinesses" 区分商户类型。\n剩余确认: 商户→CUST_ID映射(consumer_id? 还是联查DWD_CUST_MAN?)'),
    ('二、Q2(绑卡) 半解决', 'R110(完成三方支付绑卡)支持渠道筛选并标注时间范围可自定义，说明绑卡表确实存储了渠道信息。\n\n建议: 确认R110的后台源表和渠道字段名，以及"首次"的具体判定(卡级/客户级/首次绑定时间)。这一步完成后Q2可关闭。\n剩余确认: 绑卡源表名+渠道字段+首次判定逻辑。'),
    ('三、Q3(活跃阈值) 参考值已明确', 'R109(登录次数)、R117(月活跃至少6笔)、R118(微信绑卡活跃至少6笔)给出了参考阈值。\n\n建议: 0067暂时用>=1次当年登录即可(P1问题不阻塞开发)；月活跃(>=6笔)可用于手机银行深度使用指标(非0067)，未来可独立定义新指标。'),
    ('四、Q4(代发薪值域) 已解决', 'R18(是否代发客户)完整列出6种代发类型：行内代发工资、惠民惠农一卡通、省社保、市社保、国库集中支付、代发薪系统。\n\n建议: Q4关闭。0064直接用标签R18判定即可，无需逐个查CTRAKT_TYP值域。'),
    ('五、美团/微粒贷排除 直接实现', 'R125(是否美团贷款客户)和R126(是否微粒贷客户)两个标签直接可用！\n\n建议: 0062存储过程计算中，排除条件WHERE NOT R125 AND NOT R126即可。\n这是之前明确标注"待确认"的过滤条件，现已被标签表直接覆盖。'),
    ('六、0088/0063临界客户 可直接取数', 'R106(高净值临界客户)的区间定义与0063完全一致([4.5,5万)等)。\n\n建议: 0063存储过程可JOIN标签表R106进行筛选，减少重复计算。'),
    ('七、标签表与DWD/DWS源表的关系', '标签表是派生表(从DWD/DWS计算而来)，而非原始源表。\n\n建议: 存储过程主数据源仍以DWS/DWD为准，标签表用于: (a)交叉校验口径一致性 (b)提供DWD中不存在的聚合/派生数据 (c)Q1/Q2/Q7等阻塞点的快速解决方案。'),
    ('八、Q7(贵金属) 标签表无覆盖', '125个标签中未发现"贵金属持有"、"贵金属代销收入"等标签。说明贵金属数据在客户标签体系中也尚未覆盖。\n\n建议: Q7仍需由业务负责人确认源表。0065开发先做保险+理财代销收入，贵金属部分预留字段后续补充。'),
    ('九、可新引入的指标体系维度', '标签表提供了一些DWD/DWS表未直接覆盖的分析维度:\n1. 客户等级升降级(R102-R105): 可用于0052-0054的细化分析\n2. 最低权益等级(R98-R101): 多时间窗口的客户稳定性评分\n3. 交易渠道偏好(R115): 可定义新指标"手机银行活跃交易客户数"\n4. 连续代发月数(R77): 可定义新指标"稳定代发薪客户数"\n建议: 作为二阶段扩展指标纳入需求。'),
]

row = 4
for label, content in findings:
    style_cell(ws5, row, 1, label, bold_font)
    style_cell(ws5, row, 2, content)
    ws5.row_dimensions[row].height = 120
    row += 1

set_widths(ws5, [35, 95])

# Save
output = r'd:\AI\AI-Workspace\Kingbase-CRM-AI-Development-Guide\requirements\标签-指标映射分析报告.xlsx'
wb.save(output)
print(f'Saved: {output}')
