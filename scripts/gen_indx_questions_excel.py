import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

header_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
title_font = Font(name='Microsoft YaHei', bold=True, size=14, color='1F4E79')
subtitle_font = Font(name='Microsoft YaHei', bold=True, size=11, color='1F4E79')
normal_font = Font(name='Microsoft YaHei', size=10)
red_font = Font(name='Microsoft YaHei', bold=True, size=10, color='CC0000')
green_font = Font(name='Microsoft YaHei', size=10, color='006600')
row_fill_alt = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_row(ws, row, cols, is_alt=False):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.alignment = wrap_align
        cell.border = thin_border
        if is_alt:
            cell.fill = row_fill_alt

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def p0_cell(ws, row, col):
    c = ws.cell(row=row, column=col)
    c.font = red_font
    c.alignment = center_align

# ===== Sheet 1: Summary =====
ws1 = wb.active
ws1.title = '问题总览'

ws1.merge_cells('A1:I1')
ws1['A1'] = '指标数据统计(ADS_STAT_INDX_DATA) - 开发待确认问题清单'
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 30

ws1.merge_cells('A2:I2')
ws1['A2'] = '版本: v1.0 | 日期: 2026-08-04'
ws1['A2'].font = Font(name='Microsoft YaHei', size=9, color='666666')
ws1['A2'].alignment = Alignment(horizontal='center')
ws1.row_dimensions[2].height = 20

headers = ['编号', '问题摘要', '等级', '影响指标', '确认对象', '已知信息', '核心确认点', '风险', '建议批次']
for c, h in enumerate(headers, 1):
    ws1.cell(row=4, column=c, value=h)
style_header(ws1, 4, len(headers))
ws1.row_dimensions[4].height = 25

questions = [
    ['Q1', '商户小微/个体区分标识',
     'P0', '0068, 0069', '支付平台技术负责人',
     'uepp_pay_mct_info.mct_type含personage(个体)/smallBusinesses(小微)；商户与CRM客户关联字段待确认',
     '1.mct_type是否即为区分标准？\n2.商户->CUST_ID映射方式\n3.AUM/活期指商户结算存款还是关联个人AUM？',
     '低', '第二批'],
    ['Q2', '绑卡"首次绑定"判断逻辑及绑定记录表',
     'P0', '0070, 0071', '手机银行系统负责人',
     'mbk_cust_acct含acct_add_date/channel，但无三方平台类型字段(微信/支付宝/云闪付)',
     '1."首次"判定规则(A/B/C)\n2.三方平台区分字段\n3.cust_no->CUST_ID映射\n4."新增持卡"定义',
     '中', '第二批'],
    ['Q3', '手机银行登录"活跃"频次阈值',
     'P1', '0067', '手机银行业务负责人',
     'mbk_cust_log_login含lgn_date/lgn_status/lgn_chnl；当前无"活跃"定义',
     '1.最低登录次数/天数阈值\n2.是否仅统计成功登录\n3.是否排除特定渠道',
     '低', '第二批'],
    ['Q4', '代发薪签约类型值域(CTRAKT_TYP)',
     'P1', '0064', 'CRM数据仓库负责人',
     '记忆卡片标注CTRAKT_TYP=\'代发\'；实际存储格式(字符串/编码)未核实',
     '1.CTRAKT_TYP完整值域\n2.代发薪对应准确值\n3.子类型是否需区分',
     '低', '第一批'],
    ['Q5', '借贷管户(MNG_TYP=2)是否纳入目标任务',
     'P0', '全部路径B', 'CRM业务分析师',
     'DWD_CUST_MAN.MNG_TYP有1(理财)/2(借贷)；机构路径未明确是否含MNG_TYP=2',
     '1.RSV_OBJ=机构下"全部客户"是否含MNG_TYP=2\n2.MNG_TYP完整值域\n3.是否需排除非个人/休眠',
     '中', '第二批'],
    ['Q6', 'DWD_MKT_ACT_TARGT.INDX_ID存储格式',
     'P0', '全部路径A', 'CRM数据仓库负责人',
     'ADS_MKT_TASK_INDX_SUB_CMPLT.INDX_ID为VARCHAR(40)；Excel指标列用INDX_0046格式',
     '1.INDX_ID格式(INDX_0046/0046/INDX0046)\n2.表DDL/测试数据可提供否\n3.INDX_CODE格式是否一致',
     '低', '第一批'],
    ['Q7', '贵金属代销收入的源表和字段',
     'P0', '0065', '贵金属业务负责人',
     'DWD_CUST_INDV_KYC.BK_OUTER_GOLD=行外贵金属(非本行)；全量DDL搜索未发现本行贵金属交易表',
     '1.贵金属交易ODS表名+库名\n2.收入/佣金字段\n3.是否分类别\n4.无数据源则先做保险+理财？',
     '高', '第三批'],
    ['Q8', 'RSV_OBJ=机构时无效客户过滤策略',
     'P1', '全部路径B', 'CRM业务分析师',
     'v2.0.3定义"有效客户"=有下挂账户；机构"全部客户"是否按此过滤待定',
     '1.是否按有效客户口径过滤\n2.是否排除注销/休眠客户\n3.0资产客户对前端的影响',
     '低', '第二批'],
]

for i, q in enumerate(questions):
    row = 5 + i
    is_alt = (i % 2 == 1)
    for c, val in enumerate(q, 1):
        ws1.cell(row=row, column=c, value=val)
    style_row(ws1, row, len(headers), is_alt)
    if q[2] == 'P0':
        p0_cell(ws1, row, 3)
    risk_cell = ws1.cell(row=row, column=8)
    risk_cell.alignment = center_align
    if q[7] == '高':
        risk_cell.font = red_font
    batch = ws1.cell(row=row, column=9)
    batch.alignment = center_align
    ws1.row_dimensions[row].height = 90

set_col_widths(ws1, [8, 28, 8, 16, 22, 38, 42, 8, 10])

ws1.merge_cells('A14:I14')
ws1['A14'] = '等级: P0=阻塞开发  P1=影响精度 | 风险: 高=可能跨部门延迟  中=需业务决策  低=可快速确认 | 批次: 第一批=可自行验证  第二批=需确认人反馈  第三批=可能阻塞较久'
ws1['A14'].font = Font(name='Microsoft YaHei', size=9, color='666666')
ws1['A14'].alignment = Alignment(wrap_text=True)

# ===== Sheet 2: By person =====
ws2 = wb.create_sheet('按确认人分组')

ws2.merge_cells('A1:G1')
ws2['A1'] = '按确认人分组的待确认问题'
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 28

people_groups = [
    ('CRM数据仓库负责人(Q4+Q6)',
     [q for q in questions if '数据仓库' in q[4]],
     '可自行验证DDL，无需等待外部反馈'),
    ('CRM业务分析师(Q5+Q8)',
     [q for q in questions if '业务分析师' in q[4]],
     '口头沟通即可，建议Q5+Q8一起确认'),
    ('支付平台技术负责人(Q1)',
     [q for q in questions if '支付平台' in q[4]],
     '书面确认，附uepp_pay_mct_info表结构截图'),
    ('手机银行系统负责人(Q2)',
     [q for q in questions if '手机银行系统' in q[4]],
     '书面确认，提供mbk_cust_acct完整值域及映射规则'),
    ('手机银行业务负责人(Q3)',
     [q for q in questions if '手机银行业务' in q[4]],
     '口头或书面均可'),
    ('贵金属业务负责人(Q7)',
     [q for q in questions if '贵金属' in q[4]],
     '正式书面确认，可能需跨部门协调，建议优先发起'),
]

p_headers = ['编号', '问题摘要', '等级', '影响指标', '核心确认点', '风险', '备注']
row = 3

for person_name, person_qs, note in people_groups:
    if not person_qs:
        continue
    ws2.merge_cells(f'A{row}:G{row}')
    ws2.cell(row=row, column=1, value=f'■ 确认人: {person_name}')
    ws2.cell(row=row, column=1).font = subtitle_font
    ws2.cell(row=row, column=1).fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    ws2.row_dimensions[row].height = 22
    row += 1

    ws2.merge_cells(f'A{row}:G{row}')
    ws2.cell(row=row, column=1, value=f'沟通建议: {note}')
    ws2.cell(row=row, column=1).font = Font(name='Microsoft YaHei', size=9, color='666666')
    row += 1

    for c, h in enumerate(p_headers, 1):
        ws2.cell(row=row, column=c, value=h)
    style_header(ws2, row, len(p_headers))
    ws2.row_dimensions[row].height = 22
    row += 1

    for j, q in enumerate(person_qs):
        vals = [q[0], q[1], q[2], q[3], q[5], q[7], q[6]]
        is_alt = (j % 2 == 1)
        for c, val in enumerate(vals, 1):
            ws2.cell(row=row, column=c, value=val)
        style_row(ws2, row, len(p_headers), is_alt)
        if q[2] == 'P0':
            p0_cell(ws2, row, 3)
        ws2.row_dimensions[row].height = 65
        row += 1
    row += 1

set_col_widths(ws2, [8, 28, 8, 16, 50, 8, 18])

# ===== Sheet 3: Detailed =====
ws3 = wb.create_sheet('P0详细确认单')

ws3.merge_cells('A1:E1')
ws3['A1'] = 'P0级问题详细确认单（可直接转发给确认人）'
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 28

detail_items = [
    {
        'id': 'Q1',
        'title': '商户(0068/0069)与个人客户如何区分？',
        'confirm_with': '支付平台技术负责人',
        'indicators': 'INDX_0068 支付结算年度价值商户数\nINDX_0069 二维码收款个人活跃商户结算存款留存率',
        'context': '0068价值商户判定含两个维度门槛:\n  小微商户: AUM月日均>=2万或活期月日均>=1万\n  个体商户: AUM月日均>=5万或活期月日均>=2.5万\n0069同样需要区分商户类型以筛选正确的留存率计算人群。',
        'known': 'ODS表 crmdm.uepp_pay_mct_info:\n  mct_type varchar(20) - personage(个体)/smallBusinesses(小微)/company(企业)/institution(党政机关)/otherOrganizations(其他组织)\n  consumer_id varchar(40) - 客户ID\n  org_id / job_id - 机构/客户经理',
        'questions': '1. mct_type="personage"和"smallBusinesses"是否即0068/0069的区分标准？\n2. 商户如何关联CRM客户号(CUST_ID)？通过consumer_id? job_id? 还是联查其他表？\n3. 0068的AUM/活期月日均指商户关联的个人客户AUM，还是商户自有结算存款？',
        'format': 'Q1回复:\n1. 商户类型: [mct_type字段/其他字段]\n2. 商户->CUST_ID映射: [联表名+JOIN条件]\n3. AUM/活期: [关联个人客户AUM / 商户结算存款 / 关联逻辑:___]',
    },
    {
        'id': 'Q2',
        'title': '0070绑卡"首次绑定"的判断逻辑？',
        'confirm_with': '手机银行系统负责人',
        'indicators': 'INDX_0070 银行卡三方支付绑卡\nINDX_0071 银行卡三方支付绑卡率',
        'context': '0070定义: 我行银行卡首次绑定微信支付、支付宝、云闪付等三方支付的数量\n0071定义: 绑卡数/本年新增持卡客户(<=70岁)的绑定率',
        'known': 'ODS表 crmdm.mbk_cust_acct:\n  cust_no varchar(32) - 电子银行客户号\n  acct varchar(32) - 卡号\n  acct_add_chnl varchar(3) - 下挂渠道(MB=手机银行/TC=柜面)\n  acct_add_date varchar(10) - 下挂日期(YYYY-MM-DD)\n\n问题: 该表无"三方平台类型"字段，也无"是否首次"标志',
        'questions': '1. "首次绑定"判定逻辑: (A)该卡号+该平台在mbk_cust_acct首条 (B)该客户+该平台首条 (C)有其他绑卡表?\n2. 微信/支付宝/云闪付如何区分(哪个表哪个字段)?\n3. mbk_cust_acct.cust_no如何映射到CRM CUST_ID?\n4. "本年新增持卡客户<=70岁"中"新增持卡"指开户日期在当年?还是首次绑定日期?',
        'format': 'Q2回复:\n1. 首次绑定: [方案A/B/C或自定义]\n2. 绑卡记录表: [表名], 三方平台字段: [字段名+值: 微信=?,支付宝=?,云闪付=?]\n3. cust_no->CUST_ID: [联表名+JOIN条件]\n4. 新增持卡: [开户日期/首次绑定日期]',
    },
    {
        'id': 'Q5',
        'title': '目标任务客户范围是否包含借贷管户？',
        'confirm_with': 'CRM业务分析师',
        'indicators': '全部27个指标在"目标任务"路径下的客户范围',
        'context': '记忆卡片v1.1定义目标任务数据路径:\n  RSV_OBJ=机构(0): 机构下属全部客户\n  RSV_OBJ=理财经理(1): 仅理财经理名下客户(MNG_TYP=1)\n机构路径未明确是否包含借贷管户(MNG_TYP=2)',
        'known': 'DWD_CUST_MAN.MNG_TYP 包含: 1(理财管户) / 2(借贷管户)\nRSV_OBJ=理财经理 已明确用 MNG_TYP=1\nRSV_OBJ=机构 说"全部客户"但未提及是否过滤MNG_TYP',
        'questions': '1. RSV_OBJ=机构下"全部客户"是否包含MNG_TYP=2(借贷管户)?\n2. MNG_TYP完整值域?是否还有其他类型(如对公管户=3)?\n3. 是否需要排除非个人/休眠/注销客户?',
        'format': 'Q5回复:\n1. RSV_OBJ=机构的客户范围: [含MNG_TYP=1+2 / 仅MNG_TYP=1 / 自定义]\n2. MNG_TYP值域: [1=理财,2=借贷,3=___]\n3. 排除条件: [无 / 排除非个人/休眠/注销]',
    },
    {
        'id': 'Q6',
        'title': 'DWD_MKT_ACT_TARGT.INDX_ID字段类型和值域？',
        'confirm_with': 'CRM数据仓库负责人',
        'indicators': '全部27个指标在"营销活动"路径下的关联过滤',
        'context': '记忆卡片v1.1新增DWD_MKT_ACT_TARGT(活动-指标关联表),\n路径A改为按此表过滤而非产出全部28个指标。\n开发时需要知道INDX_ID的准确格式才能编写JOIN条件。',
        'known': 'ADS_MKT_TASK_INDX_SUB_CMPLT.INDX_ID 定义为 VARCHAR(40)\nExcel指标列使用 INDX_0046~0079 格式(含下划线前缀)\nADS_STAT_INDX_DATA.INDX_CODE 为 VARCHAR(100)',
        'questions': '1. INDX_ID实际存储格式: (A)INDX_0046 (B)0046 (C)INDX0046?\n2. DWD_MKT_ACT_TARGT表是否已建表?DDL和测试数据可否提供?\n3. ADS_STAT_INDX_DATA.INDX_CODE格式是否与INDX_ID一致?',
        'format': 'Q6回复:\n1. INDX_ID格式: [INDX_0046 / 0046 / 其他]\n2. DWD_MKT_ACT_TARGT: [DDL可提供/待建表], 测试数据: [有/无]\n3. INDX_CODE格式: [与INDX_ID一致/不一致,需转换: ___]',
    },
    {
        'id': 'Q7',
        'title': '贵金属代销收入的源表和字段？',
        'confirm_with': '贵金属业务负责人',
        'indicators': 'INDX_0065 代销业务收入',
        'context': 'v2.0.3: 0065 = 保险代销 + 贵金属代销 + 理财代销\n两个来源已明确:\n  保险->DWD_ACCT_INSUR.NEW_INSUR_AMT(首期保费)\n  理财->DWD_ACCT_FIN(代销大类)\n贵金属代销 -> 未确定',
        'known': 'DWD_CUST_INDV_KYC.BK_OUTER_GOLD="行外贵金属"(非本行)\n全量DDL文件搜索未发现本行贵金属代销明细表\n可能的源系统: 综合理财销售(TPS)/核心(CBS)/贵金属交易系统',
        'questions': '1. 本行贵金属代销交易记录存储在哪个ODS表(表名+库名)?\n2. 代销收入/佣金对应字段名?\n3. 是否需区分实物贵金属/积存金/账户贵金属?\n4. 如当前无数据源,0065是否先做保险+理财?',
        'format': 'Q7回复:\n1. 贵金属源表: [表名+库名 / 暂无数据源]\n2. 收入字段: [字段名+含义]\n3. 子类别: [需区分:___ / 不需区分]\n4. 暂处理方案: [全做 / 先做保险+理财,贵金属后续补充]',
    },
]

row = 3
for item in detail_items:
    # Title row
    ws3.merge_cells(f'A{row}:E{row}')
    ws3.cell(row=row, column=1, value=f'{item["id"]} - {item["title"]}')
    ws3.cell(row=row, column=1).font = Font(name='Microsoft YaHei', bold=True, size=12, color='1F4E79')
    ws3.cell(row=row, column=1).fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    ws3.row_dimensions[row].height = 25
    row += 1

    fields = [
        ('确认对象', item['confirm_with']),
        ('影响指标', item['indicators']),
        ('业务背景', item['context']),
        ('已知信息', item['known']),
        ('需确认事项', item['questions']),
        ('预期回复格式', item['format']),
    ]

    for label, content in fields:
        ws3.merge_cells(f'B{row}:E{row}')
        ws3.cell(row=row, column=1, value=label)
        ws3.cell(row=row, column=1).font = Font(name='Microsoft YaHei', bold=True, size=10)
        ws3.cell(row=row, column=1).alignment = Alignment(vertical='top')
        ws3.cell(row=row, column=1).fill = row_fill_alt
        ws3.cell(row=row, column=2, value=content)
        ws3.cell(row=row, column=2).font = Font(name='Microsoft YaHei', size=10)
        ws3.cell(row=row, column=2).alignment = wrap_align
        line_count = max(content.count('\n') + 1, 3)
        ws3.row_dimensions[row].height = max(18 * line_count, 60)
        row += 1
    row += 1

set_col_widths(ws3, [16, 40, 40, 40, 40])

# ===== Sheet 4: Resolved =====
ws4 = wb.create_sheet('已解决缺口')

ws4.merge_cells('A1:F1')
ws4['A1'] = 'P0材料缺口中已确认存在的项目'
ws4['A1'].font = title_font
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 28

r_headers = ['原缺口', '原描述', '核实结果', '文件路径', '状态', '备注']
for c, h in enumerate(r_headers, 1):
    ws4.cell(row=3, column=c, value=h)
style_header(ws4, 3, len(r_headers))

resolved = [
    ['#8', 'uepp_pay_order_info / uepp_pay_mct_info ODS DDL',
     '7个uepp DDL文件已存在', 'data_assets/ddl/ods/uepp/',
     '已确认', '含order_info(支付)+mct_info(商户)+channel_info(渠道)等'],
    ['#9-a', 'mbk_cust_log_login ODS DDL',
     'DDL已存在，字段完整', 'data_assets/ddl/ods/mbk/mbk_cust_log_login.sql',
     '已确认', '转入Q3确认活跃阈值'],
    ['#9-b', 'mbk_cust_acct ODS DDL',
     'DDL已存在，字段完整', 'data_assets/ddl/ods/mbk/mbk_cust_acct.sql',
     '已确认', '转入Q2确认首次绑定+平台区分'],
]

for i, r in enumerate(resolved):
    for c, val in enumerate(r, 1):
        ws4.cell(row=4+i, column=c, value=val)
    style_row(ws4, 4+i, len(r_headers), i%2==1)
    ws4.cell(row=4+i, column=5).font = green_font
    ws4.row_dimensions[4+i].height = 30

set_col_widths(ws4, [10, 38, 35, 45, 10, 45])

# Save
output = r'd:\AI\AI-Workspace\Kingbase-CRM-AI-Development-Guide\requirements\指标开发待确认问题清单.xlsx'
wb.save(output)
print(f'Saved: {output}')
