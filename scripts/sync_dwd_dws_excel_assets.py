"""Synchronize DWD/DWS mapping Markdown and DDL from their model workbooks."""

from __future__ import annotations

import hashlib
import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOKS = (
    ("ods_to_dwd", "dwd", "DWD明细层数据模型_CRM_ V1.0.xlsx", "ods到dwd映射.md"),
    ("dwd_to_dws", "dws", "DWS汇总层数据模型_CRM_ V1.0.xlsx", "dwd到dws映射.md"),
)
TYPE_SOURCE = ROOT / "data_assets" / "mapping" / "crmdm表结构.xlsx"
TABLE_RE = re.compile(r"^(?:DWD|DWS)_[A-Z0-9_]+$")
TYPE_RE = re.compile(r"^(?:VARCHAR2?|NUMBER|NUMERIC|DECIMAL|DATE|TIMESTAMP|CHAR|BPCHAR)(?:\([^)]*\))?$", re.I)


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("|", "&#124;").replace("\r\n", "<br>").replace("\n", "<br>")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def kingbase_type(raw_type: str) -> str:
    value = raw_type.upper().replace("VARCHAR2", "VARCHAR").replace("DECIMAL", "NUMBER").replace("NUMERIC", "NUMBER")
    return value or "VARCHAR"


def load_type_source() -> dict[str, dict[str, tuple[str, str, bool]]]:
    workbook = load_workbook(TYPE_SOURCE, data_only=True, read_only=True)
    fields: dict[str, dict[str, tuple[str, str, bool]]] = {}
    for row in workbook.active.iter_rows(min_row=2, values_only=True):
        values = list(row) + [None] * 10
        table, field = clean(values[1]).upper(), clean(values[4]).upper()
        if not TABLE_RE.fullmatch(table) or not field:
            continue
        type_name = clean(values[5]).upper()
        length, scale = clean(values[6]), clean(values[7])
        nullable = clean(values[8]).upper() != "N"
        if type_name in {"VARCHAR", "VARCHAR2", "CHAR"} and length:
            field_type = f"{type_name}({length})"
        elif type_name in {"NUMBER", "NUMERIC", "DECIMAL"} and length not in {"", "0"} and scale not in {"", "0"}:
            field_type = f"{type_name}({length},{scale})"
        else:
            field_type = type_name
        fields.setdefault(table, {})[field] = (field_type, clean(values[9]), nullable)
    return fields


def table_name(worksheet) -> str:
    for row in worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, 4), values_only=True):
        for value in row:
            candidate = clean(value).upper()
            if TABLE_RE.fullmatch(candidate):
                return candidate
    return ""


def table_comment(worksheet) -> str:
    for row in worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, 2), values_only=True):
        values = list(row)
        for index, value in enumerate(values[:-1]):
            if clean(value) in {"实体中文名", "实体中文名称"}:
                return clean(values[index + 1])
    return clean(worksheet.title)


def standard_rows(worksheet) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    for row in worksheet.iter_rows(min_row=6, values_only=True):
        values = list(row) + [None] * 16
        field = clean(values[1]).upper()
        raw_type = clean(values[3]) if TYPE_RE.fullmatch(clean(values[3])) else clean(values[2])
        if not re.fullmatch(r"[A-Z_][A-Z0-9_$]*", field) or not TYPE_RE.fullmatch(raw_type):
            if records:
                break
            continue
        records.append({
            "field": field,
            "type": raw_type,
            "comment": clean(values[4]),
            "source_table": clean(values[11]),
            "source_field": clean(values[13]),
            "rule": clean(values[15]),
        })
    return records


def column_a_rows(worksheet, target_table: str) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        values = list(row) + [None] * 18
        field = clean(values[0]).upper()
        if field == "变更登记":
            break
        if not re.fullmatch(r"[A-Z_][A-Z0-9_$]*", field):
            continue
        raw_type = clean(values[1])
        if not TYPE_RE.fullmatch(raw_type):
            continue
        records.append({
            "field": field,
            "type": raw_type,
            "comment": clean(values[2]),
            "source_table": clean(values[5]),
            "source_field": clean(values[7]),
            "rule": clean(values[9]),
        })
    return records


def block_rows(worksheet, target_table: str) -> list[dict[str, str]]:
    records: OrderedDict[str, dict[str, str]] = OrderedDict()
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        values = list(row) + [None] * 11
        if clean(values[0]).upper() != target_table:
            continue
        field = clean(values[1]).upper()
        if not re.fullmatch(r"[A-Z_][A-Z0-9_$]*", field) or field in records:
            continue
        records[field] = {
            "field": field,
            "type": "",
            "comment": clean(values[2]),
            "source_table": clean(values[5]),
            "source_field": clean(values[7]),
            "rule": clean(values[9]),
        }
    return list(records.values())


def extract_workbook(path: Path, type_source: dict[str, dict[str, tuple[str, str, bool]]]) -> OrderedDict[str, dict]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    tables: OrderedDict[str, dict] = OrderedDict()
    for worksheet in workbook.worksheets:
        target = table_name(worksheet)
        if not target:
            continue
        records = standard_rows(worksheet)
        if not records:
            records = column_a_rows(worksheet, target)
        if not records:
            records = block_rows(worksheet, target)
        if not records:
            continue
        existing_fields = {record["field"] for record in records}
        workbook_values = {
            clean(cell.value).upper()
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        }
        for field, fallback in type_source.get(target, {}).items():
            if field in workbook_values and field not in existing_fields:
                records.append({
                    "field": field,
                    "type": fallback[0],
                    "comment": fallback[1],
                    "source_table": "",
                    "source_field": "",
                    "rule": "",
                })
        for record in records:
            fallback = type_source.get(target, {}).get(record["field"])
            if not record["type"] and fallback:
                record["type"] = fallback[0]
            if not record["comment"] and fallback:
                record["comment"] = fallback[1]
            if not record["type"]:
                raise ValueError(f"{target}.{record['field']} 缺少字段类型")
        tables[target] = {"comment": table_comment(worksheet), "fields": records}
    return tables


def existing_not_null(path: Path) -> set[str]:
    if not path.exists():
        return set()
    return {
        match.group(1).upper()
        for match in re.finditer(r"^\s*([A-Za-z_][A-Za-z0-9_$]*)\s+[^,\n]+\bNOT\s+NULL\b", path.read_text(encoding="utf-8", errors="replace"), re.I | re.M)
    }


def render_ddl(table: str, definition: dict, preserved_not_null: set[str], type_source: dict[str, dict[str, tuple[str, str, bool]]]) -> str:
    lines = ["/*", f" * {table}", f" * 中文名称: {definition['comment']}", " * 来源: Mapping Excel 同步", " */", "", f"CREATE TABLE IF NOT EXISTS {table} ("]
    columns = []
    field_count = len(definition["fields"])
    for index, record in enumerate(definition["fields"]):
        field = record["field"]
        source = type_source.get(table, {}).get(field)
        required = field in preserved_not_null or (source is not None and not source[2])
        nullable = " NOT NULL" if required else " NULL"
        separator = "," if index < field_count - 1 else ""
        columns.append(f"    {field} {kingbase_type(record['type'])}{nullable}{separator} -- {record['comment']}")
    lines.extend(columns)
    lines.extend([");", "", f"COMMENT ON TABLE {table} IS '{definition['comment'].replace("'", "''")}';",
    ])
    for record in definition["fields"]:
        lines.append(f"COMMENT ON COLUMN {table}.{record['field']} IS '{record['comment'].replace("'", "''")}';")
    return "\n".join(lines) + "\n"


def render_mapping(flow: str, workbook: Path, tables: OrderedDict[str, dict]) -> str:
    lines = [f"# {flow.replace('_to_', '到').upper()}映射 字段映射", "", "## 映射来源", "", f"- Excel：`{workbook.relative_to(ROOT).as_posix()}`", f"- Excel SHA-256：`{sha256(workbook)}`", "", "## 映射概览", "", "| 目标表 | 字段数 |", "|--------|-------:|"]
    lines.extend(f"| {table} | {len(definition['fields'])} |" for table, definition in tables.items())
    lines.extend(["", "## 字段映射详情", ""])
    for table, definition in tables.items():
        lines.extend([f"### {table}", "", "| 目标字段 | 目标字段中文名 | 目标字段类型 | 源表 | 源字段 | 映射规则 |", "|----------|----------------|--------------|------|--------|----------|"])
        for record in definition["fields"]:
            lines.append(f"| {record['field']} | {record['comment']} | {record['type']} | {record['source_table']} | {record['source_field']} | {record['rule']} |")
        lines.append("")
    lines.extend(["---", "", "*本文件由对应 Excel 模型同步生成；Excel 更新后必须重新生成本文件。", ""])
    return "\n".join(lines)


def main() -> None:
    type_source = load_type_source()
    for flow, layer, workbook_name, mapping_name in WORKBOOKS:
        workbook = ROOT / "data_assets" / "mapping" / flow / workbook_name
        mapping_path = workbook.parent / mapping_name
        tables = extract_workbook(workbook, type_source)
        mapping_path.write_text(render_mapping(flow, workbook, tables), encoding="utf-8", newline="\n")
        ddl_dir = ROOT / "data_assets" / "ddl" / layer
        for table, definition in tables.items():
            path = ddl_dir / f"{table.lower()}.sql"
            path.write_text(render_ddl(table, definition, existing_not_null(path), type_source), encoding="utf-8", newline="\n")
        print(f"{flow}: Mapping={mapping_path.relative_to(ROOT)}, DDL={len(tables)} 张")


if __name__ == "__main__":
    main()
