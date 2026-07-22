"""Apply the approved 7-character organization-id contract to DWD/DWS/ADS assets."""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOKS = (
    ROOT / "data_assets/mapping/ods_to_dwd/DWD明细层数据模型_CRM_ V1.0.xlsx",
    ROOT / "data_assets/mapping/dwd_to_dws/DWS汇总层数据模型_CRM_ V1.0.xlsx",
    ROOT / "data_assets/mapping/dws_to_ads/ADS应用层数据模型_CRM_ V1.0.xlsx",
)
ORG_FIELDS = {
    "ORG_ID", "ORG_LEAD", "SUP_ORG_ID", "OPEN_ACCT_ORG", "OPRT_ORG", "MKT_ORG",
    "CREAT_ORG", "TX_ORG", "ACCT_BLNG_ORG", "COSPSR_ORG", "MKT_PERSN_ORG",
    "BLNG_BRCH", "BLNG_BRCH_SUB", "BLNG_BRCH_NET",
}


def update_workbook(path: Path) -> int:
    workbook = openpyxl.load_workbook(path)
    changes = 0
    for sheet in workbook.worksheets:
        for row in range(1, sheet.max_row + 1):
            field = str(sheet.cell(row, 2).value or "").strip().upper()
            if field in ORG_FIELDS:
                for column in (3, 4):
                    value = sheet.cell(row, column).value
                    if isinstance(value, str) and re.fullmatch(r"VARCHAR(?:2|7|20)?\(\d+\)", value.strip(), re.I):
                        sheet.cell(row, column).value = "VARCHAR2(7)"
                        changes += 1
                if str(sheet.cell(row, 3).value or "").strip().upper() in {"VARCHAR", "VARCHAR2"} and sheet.cell(row, 4).value != 7:
                    sheet.cell(row, 4).value = 7
                    changes += 1
            if field == "STATIS_OBJ":
                for column in (3, 4):
                    value = sheet.cell(row, column).value
                    if isinstance(value, str) and re.fullmatch(r"VARCHAR(?:2|7|20)?\(\d+\)", value.strip(), re.I):
                        sheet.cell(row, column).value = "VARCHAR2(20)"
                        changes += 1
                if str(sheet.cell(row, 3).value or "").strip().upper() in {"VARCHAR", "VARCHAR2"} and sheet.cell(row, 4).value != 20:
                    sheet.cell(row, 4).value = 20
                    changes += 1
    workbook.save(path)
    return changes


def update_text_assets() -> int:
    changes = 0
    for relative in ("data_assets/ddl/dwd", "data_assets/ddl/dws", "data_assets/ddl/ads", "data_assets/data_dictionary/dwd", "data_assets/data_dictionary/dws", "data_assets/data_dictionary/ads"):
        for path in (ROOT / relative).glob("*"):
            if path.suffix not in {".sql", ".md"}:
                continue
            content = path.read_text(encoding="utf-8")
            original = content
            for field in ORG_FIELDS:
                content = re.sub(rf"(\b{field}\b[^\n]*?VARCHAR2?\()\d+(\))", r"\g<1>7\2", content, flags=re.I)
                content = re.sub(rf"(\|\s*{field}\s*\|[^\n]*?VARCHAR2?\(?)(?:\d+)(\)?\s*\|)", r"\g<1>7\2", content, flags=re.I)
            content = re.sub(r"(\bSTATIS_OBJ\b[^\n]*?VARCHAR2?\()\d+(\))", r"\g<1>20\2", content, flags=re.I)
            content = re.sub(r"(\|\s*STATIS_OBJ\s*\|[^\n]*?VARCHAR2?\s*\|\s*)\d+", r"\g<1>20", content, flags=re.I)
            if content != original:
                path.write_text(content, encoding="utf-8")
                changes += 1
    return changes


def main() -> None:
    workbook_changes = sum(update_workbook(path) for path in WORKBOOKS)
    text_changes = update_text_assets()
    print(f"已更新 {workbook_changes} 个模型字段和 {text_changes} 个 DDL/数据字典文件。")


if __name__ == "__main__":
    main()
