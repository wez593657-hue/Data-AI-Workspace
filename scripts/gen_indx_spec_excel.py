import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# === Style definitions ===
hdr_font = Font(name='Microsoft YaHei', bold=True, size=10, color='FFFFFF')
hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
cat_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
section_font = Font(name='Microsoft YaHei', bold=True, size=11, color='1F4E79')
normal_font = Font(name='Microsoft YaHei', size=10)
bold_font = Font(name='Microsoft YaHei', bold=True, size=10)
green_font = Font(name='Microsoft YaHei', size=10, color='006600')
orange_font = Font(name='Microsoft YaHei', size=10, color='BF8F00')
red_font = Font(name='Microsoft YaHei', bold=True, size=10, color='CC0000')
confirmed_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
pending_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
alt_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_top = Alignment(wrap_text=True, vertical='top')
wrap_center = Alignment(wrap_text=True, vertical='top', horizontal='center')

def style_hdr(ws, r, cols):
    for c in range(1, cols+1):
        cl = ws.cell(row=r, column=c)
        cl.font = hdr_font; cl.fill = hdr_fill; cl.alignment = wrap_center; cl.border = thin_border

def style_cell(ws, r, c, val, font=None, fill=None, align=None):
    cl = ws.cell(row=r, column=c, value=val)
    cl.font = font or normal_font
    cl.alignment = align or wrap_top
    cl.border = thin_border
    if fill: cl.fill = fill

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def section_row(ws, r, cols, text, fill=cat_fill):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    style_cell(ws, r, 1, text, section_font, fill)
    ws.row_dimensions[r].height = 24

def kv_row(ws, r, cols, label, value, label_w=1):
    ws.merge_cells(start_row=r, start_column=label_w+1, end_row=r, end_column=cols)
    style_cell(ws, r, label_w, label, bold_font)
    style_cell(ws, r, label_w+1, value)

# ================================================================
# Indicator Data
# ================================================================

DIMENSIONS_DETAIL = (
    '■ 统计口径维度(STATIS_CALIB)\n'
    '  取值: "营销活动" / "目标任务"\n'
    '  对应: ADS_STAT_INDX_DATA.STATIS_CALIB\n'
    '\n'
    '■ 数据归属维度(DATA_BLNG)\n'
    '  取值: 机构编号 / 客户经理编号\n'
    '  对应: ADS_STAT_INDX_DATA.DATA_BLNG\n'
    '\n'
    '■ 统计对象维度(STATIS_DIM)\n'
    '  路径A: MKT_ACT_ID\n'
    '  路径B: TSK_ID\n'
    '  对应: ADS_STAT_INDX_DATA.STATIS_DIM\n'
    '\n'
    '■ 时间维度\n'
    '  DATA_DATE = 跑批日期(V_SYSDAT, 格式YYYYMMDD)\n'
    '  期初基准: ACT_BGN_DATE(DWD_MKT_ACT_INFO) / TSK_BGN_DATE(DWD_MKT_INDX_TSK)\n'
    '  期末: DATA_DATE\n'
    '  年初基准: 当年1月1日(DWS_CUST_ASSE_LIAB, DATA_DATE)\n'
    '  上月末: 上月最后一天(DWS_CUST_ASSE_LIAB, DATA_DATE)\n'
    '  上季末: 上季最后一天(DWS_CUST_ASSE_LIAB, DATA_DATE)\n'
    '  统计周期: [ACT_BGN_DATE/TSK_BGN_DATE, DATA_DATE] 每日\n'
    '  对应表: ADS_STAT_INDX_DATA.DATA_DATE\n'
    '\n'
    '■ 客户维度\n'
    '  客户号: CUST_ID\n'
    '  法人行号: PERSN_LEGAL_BK_CODE\n'
    '  数据源一 DWS_CUST_ASSE_LIAB: BAL_TYPE=1(日)/2(月日均)/4(年日均)\n'
    '  数据源二 DWD_CUST_MAN: MNG_TYP=1(理财管户, 仅MNG_TYP=1v1.2决议)\n'
    '  数据源三 DWD_CUST_LVL_INFO: CUST_LVL(0-7级客户等级)\n'
    '  数据源四 DWD_CUST_INDV_INFO: BIRTH_DATE(年龄), OPEN_DATE(开户日期)\n'
    '\n'
    '■ 机构维度(递归向上汇总)\n'
    '  数据源 DWD_SYS_ORG: ORG_ID → SUP_ORG_ID 递归\n'
    '  路径A: DWD_MKT_ACT_ORG_REL.PRTSPT_ORG → 按SUP_ORG_ID汇总上级机构\n'
    '  路径B: RSV_OBJ=机构时, DWD_MKT_INDX_TSK.RSV_OBJ_ID → 递归获取下级所有机构\n'
    '         从机构获取客户: DWD_CUST_MAN.MNG_TYP=1, 通过机构→职位→客户经理→客户\n'
    '\n'
    '■ 客户经理维度\n'
    '  职位-机构关联 DWD_SYS_POST: POST_ID, ORG_ID, JOB_CLS(岗位类别)\n'
    '  客户-经理管户 DWD_CUST_MAN: MNGR_POST_ID=POST_ID, MNG_TYP=1\n'
    '  路径A: DWD_MKT_TSK_INFO.MKT_PERSN → DWD_SYS_POST.POST_ID → 经理所在机构\n'
    '  路径B: DWD_MKT_INDX_TSK.RSV_OBJ=理财经理, RSV_OBJ_ID=经理编号\n'
    '\n'
    '■ 指标维度(路径A特有)\n'
    '  活动-指标关联: DWD_MKT_ACT_TARGT(MKT_ACT_ID, INDX_ID=INDX_0046格式)\n'
    '  活动类型过滤: DWD_MKT_ACT_INFO.ACT_TYP IN ("综合方案","价值提升")\n'
    '\n'
    '■ 维度展开规则\n'
    '  机构行: 机构下属所有理财管户客户(MNG_TYP=1)汇总 → DATA_BLNG=机构编号\n'
    '  经理行: 每个理财经理名下客户分别汇总 → DATA_BLNG=POST_ID\n'
    '  路径A产出: PRTSPT_ORG(汇总行)+MKT_PERSN(客户经理行)\n'
    '  路径B产出: RSV_OBJ=机构→机构行+下属经理行; RSV_OBJ=理财经理→CREAT_ORG+经理行'
)

dimensions_short = '见"维度划分详情"Sheet'
indicators = [
    # ---- Savings (0046-0051) ----
    {
        'code': '0046', 'name': '储蓄存款余额较年初', 'category': '储蓄存款',
        'unit': '元', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '衡量考核期内客户储蓄存款余额相对年初基数的增长情况，反映储蓄业务在年度维度的净增量。用于评估全年储蓄业务发展成效。',
        'biz_use': '营销活动考核指标|目标任务完成率评估|支行/客户经理年度储蓄业绩排名',
        'calc_rule': 'CURNT_VAL = SUM(本期AUM - 年初AUM)\n【v1.3】储蓄存款直接取DWS_CUST_ASSE_LIAB.AUM_BAL',
        'source_tables': 'DWS_CUST_ASSE_LIAB (BAL_TYPE=1, AUM_BAL字段)\n年初值: DWS_CUST_ASSE_LIAB (DATA_DATE=当年1月1日)',
        'stat_cycle': '跑批日当天快照 vs 当年1月1日快照',
        'dimensions': '',
        'special_cases': '年初基准日如为非交易日，取最近一个交易日数据\n客户同时参与多个活动时，各活动独立统计',
        'confirmed': [
            '[口径v1.3] 储蓄存款直接取AUM_BAL',
            '[源表] DWS_CUST_ASSE_LIAB, BAL_TYPE=1, AUM_BAL字段',
            '[基准] 对比年初(1月1日)DWS快照',
        ],
        'pending': [



        ],
        'version': 'v1.2记忆卡片',
    },
    {
        'code': '0047', 'name': '储蓄存款余额较基数', 'category': '储蓄存款',
        'unit': '元', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '衡量客户储蓄存款余额相对业务导入基数的变化。基数由Excel模板导入，通常代表营销活动启动前或目标任务下达时的初始值。',
        'biz_use': '营销活动中作为自定义基准的增量考核|目标任务分配时的个性化指标',
        'calc_rule': 'CURNT_VAL = SUM(本期AUM - BASE_VAL)\nBASE_VAL = DWD_MKT_TSK_INFO.BASE_VAL(任务级基数)\n【v1.3】储蓄存款 = AUM_BAL',
        'source_tables': 'DWS_CUST_ASSE_LIAB (BAL_TYPE=1, AUM_BAL字段) + DWD_MKT_TSK_INFO (BASE_VAL字段)',
        'stat_cycle': '跑批日当天快照 vs 业务导入基数',
        'dimensions': '',
        'special_cases': '若BASE_VAL为NULL，该指标值为NULL（不参与统计）\n基数由业务用户导入，非系统自动计算',
        'confirmed': [
            '[口径] 较基数=本期储蓄余额-业务导入BASE_VAL',
            '[源表] DWD_MKT_TSK_INFO.BASE_VAL',
            '[基数] 业务导入，非系统计算',
        ],
        'pending': [



        ],
        'version': 'v1.2记忆卡片',
    },
    {
        'code': '0048', 'name': '储蓄存款余额较上月', 'category': '储蓄存款',
        'unit': '元', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '衡量客户储蓄存款较上月末的月度增量，反映短期储蓄业务发展趋势。',
        'biz_use': '月度营销活动进度追踪|月度目标任务完成监控',
        'calc_rule': 'CURNT_VAL = SUM(本期AUM - 上月末AUM)\n上月末: DWS_CUST_ASSE_LIAB, DATA_DATE=上月最后一天\n【v1.3】AUM_BAL直接取值',
        'source_tables': 'DWS_CUST_ASSE_LIAB (BAL_TYPE=1, AUM_BAL字段, 当日+上月最后一天快照)',
        'stat_cycle': '跑批日 vs 上月最后一天',
        'dimensions': '',
        'special_cases': '月首日: 上月最后一天 = T-1日',
        'confirmed': [
            '[口径] 较上月=本期-上月末余额',
            '[源表] DWS_CUST_ASSE_LIAB',
            '[基准] 上月最后一天快照',
        ],
        'pending': [



        ],
        'version': 'v1.2记忆卡片',
    },
    {
        'code': '0049', 'name': '储蓄存款余额较上季', 'category': '储蓄存款',
        'unit': '元', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '衡量客户储蓄存款较上季末的季度增量，反映中长期储蓄业务变化趋势。',
        'biz_use': '季度营销活动考核|季度经营分析',
        'calc_rule': 'CURNT_VAL = SUM(本期AUM - 上季末AUM)\n上季末: DWS_CUST_ASSE_LIAB, DATA_DATE=上季最后一天\n【v1.3】AUM_BAL直接取值',
        'source_tables': 'DWS_CUST_ASSE_LIAB (BAL_TYPE=1, AUM_BAL字段, 当日+上季最后一天快照)',
        'stat_cycle': '跑批日 vs 上季最后一天',
        'dimensions': '',
        'special_cases': '季首日: 上季最后一天 = T-1日',
        'confirmed': [
            '[口径] 较上季=本期-上季末余额',
            '[源表] DWS_CUST_ASSE_LIAB',
            '[基准] 上季最后一天快照',
        ],
        'pending': [



        ],
        'version': 'v1.2记忆卡片',
    },
    {
        'code': '0050', 'name': '储蓄存款年日均净增额', 'category': '储蓄存款',
        'unit': '元', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '当年储蓄存款每日余额的算术平均值减去上年年日均，反映年度平均储蓄水平的变化。消除单日波动，更准确衡量全年储蓄经营水平。',
        'biz_use': '年度经营考核|客户经理年度储蓄业务评价',
        'calc_rule': 'CURNT_VAL = 当年年日均AUM - 上年年日均AUM\n年日均 = SUM(1月1日至跑批日每日AUM) / 已过天数\n【v1.3】AUM_BAL直接取值',
        'source_tables': 'DWS_CUST_ASSE_LIAB (BAL_TYPE=4, AUM_BAL字段)\n上年年日均: DWS_CUST_ASSE_LIAB (上年最后一天快照)',
        'stat_cycle': '年初至跑批日累计 / 已过天数',
        'dimensions': '',
        'special_cases': '跨年时上年年日均取上年最后一天累计值\n若跑批日=1月1日，已过天数=1',
        'confirmed': [
            '[口径] 当年年日均-上年年日均',
            '[源表] DWS_CUST_ASSE_LIAB, BAL_TYPE=4',
            '[计算] 累计余额/已过天数',
        ],
        'pending': [



        ],
        'version': 'v1.2记忆卡片',
    },
    {
        'code': '0051', 'name': '储蓄存款月日均净增额', 'category': '储蓄存款',
        'unit': '元', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '当月储蓄存款每日余额平均值减去上月月日均，反映月度平均储蓄水平的环比变化。',
        'biz_use': '月度经营分析|月度考核',
        'calc_rule': 'CURNT_VAL = 当月月日均AUM - 上月月日均AUM\n月日均 = SUM(当月1日至跑批日每日AUM) / 当月已过天数\n【v1.3】AUM_BAL直接取值',
        'source_tables': 'DWS_CUST_ASSE_LIAB (BAL_TYPE=2, AUM_BAL字段)\n上月月日均: DWS_CUST_ASSE_LIAB (上月最后一天快照)',
        'stat_cycle': '月初至跑批日累计 / 当月已过天数',
        'dimensions': '',
        'special_cases': '月首日V_SYSDAT=当月1日时，当月已过天数=1\n上月月日均取上月最后一天累计快照',
        'confirmed': [
            '[口径] 当月月日均-上月月日均',
            '[源表] DWS_CUST_ASSE_LIAB, BAL_TYPE=2',
            '[计算] 累计余额/当月已过天数',
        ],
        'pending': [



        ],
        'version': 'v1.2记忆卡片',
    },

    # ---- Customer tier (0052-0054) ----
    {
        'code': '0052', 'name': '财富及以上层级客户数增量', 'category': '客户层级',
        'unit': '户', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '活动/任务期间客户等级从财富以下(LVL<4)上升到财富及以上(LVL>=4)的净增客户数。反映客户价值提升效果。',
        'biz_use': '客户价值提升考核|营销活动升级转化率评估',
        'calc_rule': 'CURNT_VAL = COUNT(期初LVL<4 AND 期末LVL>=4)\n期初 = 活动开始日/任务下达日\n期末 = 跑批日期\n客户等级: 0=睡眠户,1=大众,2=潜力,3=优质,4=财富1,5=财富2,6=贵宾,7=私行\n(等级判断按AUM月日均而非当前余额)',
        'source_tables': 'DWS_CUST_LVL_INFO (客户等级表)',
        'stat_cycle': '期初 → 跑批日累计',
        'dimensions': '',
        'special_cases': '等级跳级(如3->6)也算作财富及以上增量(满足期末>=4即有1次)\n降级客户从LVL>=4降到<4不扣减(只计净增不下减)\n多个活动重叠时各自独立统计',
        'confirmed': [
            '[口径] 期初<4->期末>=4的客户数',
            '[源表] DWS_CUST_LVL_INFO',
            '[等级] 0-7级,基于AUM月日均判断',
        ],
        'pending': [],
        'version': 'v1.2记忆卡片',
    },
    {
        'code': '0053', 'name': '贵宾及以上层级客户数增量', 'category': '客户层级',
        'unit': '户', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '活动/任务期间客户等级从贵宾以下(LVL<6)上升到贵宾及以上(LVL>=6)的净增客户数。',
        'biz_use': '高价值客户升级考核',
        'calc_rule': 'CURNT_VAL = COUNT(期初LVL<6 AND 期末LVL>=6)',
        'source_tables': 'DWS_CUST_LVL_INFO',
        'stat_cycle': '同0052',
        'dimensions': '',
        'special_cases': '同0052逻辑',
        'confirmed': [
            '[口径] 期初<6->期末>=6的客户数',
            '[源表] DWS_CUST_LVL_INFO',
        ],
        'pending': [],
        'version': 'v1.2记忆卡片',
    },
    {
        'code': '0054', 'name': '私行及以上层级客户数增量', 'category': '客户层级',
        'unit': '户', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '活动/任务期间客户等级从私行以下(LVL<7)上升到私行(LVL=7)的净增客户数。',
        'biz_use': '最高等级客户增长考核',
        'calc_rule': 'CURNT_VAL = COUNT(期初LVL<7 AND 期末LVL=7)',
        'source_tables': 'DWS_CUST_LVL_INFO',
        'stat_cycle': '同0052',
        'dimensions': '',
        'special_cases': '同0052逻辑',
        'confirmed': [
            '[口径] 期初<7->期末=7的客户数',
            '[源表] DWS_CUST_LVL_INFO',
        ],
        'pending': [],
        'version': 'v1.2记忆卡片',
    },

    # ---- WM products (0055-0060) ----
    {
        'code': '0055', 'name': '个人理财产品年日均净增额', 'category': '理财产品',
        'unit': '元', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '当年每日个人理财产品余额(自营+代销)的算术平均值减去上年年日均。',
        'biz_use': '年度理财业务考核',
        'calc_rule': 'CURNT_VAL = 当年年日均FIN_BAL - 上年年日均FIN_BAL\nFIN_BAL取自DWS_CUST_ASSE_LIAB, BAL_TYPE=4\n【v1.3】全部理财(自营+代销, PRDKT_CATE_BIG 1/2/3/4均含)',
        'source_tables': 'DWS_CUST_ASSE_LIAB (BAL_TYPE=4)\n上年: DWS_CUST_ASSE_LIAB_CUMU_HIS\n【v1.3确认】全量无PRDKT_CATE_BIG过滤',
        'stat_cycle': '年初至跑批日累计/已过天数',
        'dimensions': '',
        'special_cases': '同0050日均计算逻辑，全量理财',
        'confirmed': [
            '[口径] 当年年日均FIN_BAL-上年年日均FIN_BAL',
            '[源表] DWS_CUST_ASSE_LIAB, BAL_TYPE=4',
            '[v1.3] 全量无过滤',
        ],
        'pending': [

        ],
        'version': 'v1.3记忆卡片',
    },
    {
        'code': '0056', 'name': '个人理财产品月日均净增额', 'category': '理财产品',
        'unit': '元', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '当月每日理财产品余额均值减去上月月日均。',
        'biz_use': '月度理财业务考核',
        'calc_rule': 'CURNT_VAL = 当月月日均FIN_BAL - 上月月日均FIN_BAL',
        'source_tables': 'DWS_CUST_ASSE_LIAB (BAL_TYPE=2)',
        'stat_cycle': '同0051',
        'dimensions': '',
        'special_cases': '同0051',
        'confirmed': [
            '[口径] 当月月日均-上月月日均',
            '[源表] DWS_CUST_ASSE_LIAB, BAL_TYPE=2',
        ],
        'pending': [

        ],
        'version': 'v1.2记忆卡片',
    },
    {
        'code': '0057', 'name': '个人理财产品销量净增额', 'category': '理财产品',
        'unit': '元', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '活动/任务期间客户购买理财产品金额减去赎回金额的净增量。基于账户明细表的实际交易流水计算。',
        'biz_use': '理财产品销售能力考核',
        'calc_rule': 'CURNT_VAL = SUM(期间购买金额) - SUM(期间赎回金额)\n赎回判定: DWD_ACCT_FIN中该产品余额变为0的记录即为赎回\n【v1.3】0057=全部理财(PRDKT_CATE_BIG 1/2/3/4均含)，0058-0060=仅代销(IN("1","2"))',
        'source_tables': 'DWD_ACCT_FIN (理财账户信息表)\n【v1.3确认】0057无PRDKT_CATE_BIG过滤，全量统计',
        'stat_cycle': '活动/任务开始日至跑批日累计',
        'dimensions': '',
        'special_cases': '部分赎回(余额减少但未到0)暂无法精确拆分，以余额归0为赎回判定\n同一天有买有卖的客户按净额计算',
        'confirmed': [
            '[口径] 期间购买-赎回(余额归0=赎回)',
            '[源表] DWD_ACCT_FIN',
            '[过滤] PRDKT_CATE_BIG NOT IN(\'1\',\'3\')',
        ],
        'pending': [

        ],
        'version': 'v1.2记忆卡片',
    },
    {
        'code': '0058', 'name': '个人代销理财产品年日均净增额', 'category': '代销理财',
        'unit': '元', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '当年每日代销理财产品余额均值减去上年年日均。仅统计代销类产品。',
        'biz_use': '年度代销业务考核',
        'calc_rule': 'CURNT_VAL = 当年年日均代销FIN_BAL - 上年年日均代销FIN_BAL\n【v1.3确认】PRDKT_CATE_BIG IN ("1","2") = 代销',
        'source_tables': 'DWS_CUST_ASSE_LIAB (BAL_TYPE=4)\n过滤: PRDKT_CATE_BIG IN ("1","2")',
        'stat_cycle': '同0055',
        'dimensions': '',
        'special_cases': '同0055',
        'confirmed': [
            '[口径] 当年年日均-上年年日均,仅代销',
            '[源表] DWS_CUST_ASSE_LIAB, BAL_TYPE=4',
            '[v1.3确认] PRDKT_CATE_BIG IN("1","2")',
        ],
        'pending': [

        ],
        'version': 'v1.3记忆卡片',
    },
    {
        'code': '0059', 'name': '个人代销理财产品月日均净增额', 'category': '代销理财',
        'unit': '元', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '当月每日代销理财产品余额均值减去上月月日均。',
        'biz_use': '月度代销业务考核',
        'calc_rule': 'CURNT_VAL = 当月月日均代销FIN_BAL - 上月月日均代销FIN_BAL\n【v1.3确认】PRDKT_CATE_BIG IN ("1","2") = 代销',
        'source_tables': 'DWS_CUST_ASSE_LIAB (BAL_TYPE=2)\n过滤: PRDKT_CATE_BIG IN ("1","2")',
        'stat_cycle': '同0056',
        'dimensions': '',
        'special_cases': '同0056',
        'confirmed': [
            '[口径] 当月月日均-上月月日均,仅代销',
            '[源表] DWS_CUST_ASSE_LIAB, BAL_TYPE=2',
            '[v1.3确认] PRDKT_CATE_BIG IN("1","2")',
        ],
        'pending': [

        ],
        'version': 'v1.3记忆卡片',
    },
    {
        'code': '0060', 'name': '个人代销理财产品销量净增额', 'category': '代销理财',
        'unit': '元', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '活动/任务期间代销理财购买-赎回净额。',
        'biz_use': '代销产品销售能力考核',
        'calc_rule': 'CURNT_VAL = SUM(期间代销理财购买) - SUM(期间代销赎回)\n赎回判定同0057\n【v1.3确认】PRDKT_CATE_BIG IN ("1","2") = 代销',
        'source_tables': 'DWD_ACCT_FIN\n过滤: PRDKT_CATE_BIG IN ("1","2")',
        'stat_cycle': '同0057',
        'dimensions': '',
        'special_cases': '同0057',
        'confirmed': [
            '[口径] 期间购买-赎回,仅代销大类',
            '[源表] DWD_ACCT_FIN',
            '[v1.3确认] PRDKT_CATE_BIG IN("1","2")',
        ],
        'pending': [

        ],
        'version': 'v1.3记忆卡片',
    },

    # ---- Insurance / Loan (0061-0062) ----
    {
        'code': '0061', 'name': '代销保险保费净增额', 'category': '保险/贷款',
        'unit': '元', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '活动/任务期间客户新投保的首期保费总额(扣除已退保保单)。反映保险代销业务新增保费收入。',
        'biz_use': '保险代销业绩考核',
        'calc_rule': 'CURNT_VAL = SUM(NEW_INSUR_AMT)\nWHERE CANCL_INSUR_DATE IS NULL (未退保)\nAND LAST_TX_DATE BETWEEN 活动/任务开始日 AND 跑批日\n(LAST_TX_DATE>=期初 AND <=跑批日)',
        'source_tables': 'DWD_ACCT_INSUR (保险账户信息)\n字段: NEW_INSUR_AMT(首期保费), CANCL_INSUR_DATE(退保日期), LAST_TX_DATE(最近交易日)',
        'stat_cycle': '期初至跑批日累计',
        'dimensions': '',
        'special_cases': '已退保保单(CANCL_INSUR_DATE IS NOT NULL)不计入\n续期保费不计入(仅NEW_INSUR_AMT即首期)\nLAST_TX_DATE在考核期内的保单才计入(保险唤醒判定同期逻辑)',
        'confirmed': [
            '[口径] 首期保费,扣除退保',
            '[源表] DWD_ACCT_INSUR',
            '[字段] NEW_INSUR_AMT, CANCL_INSUR_DATE IS NULL',
            '[时间] LAST_TX_DATE在考核区间内',
        ],
        'pending': [

        ],
        'version': 'v1.2记忆卡片',
    },
    {
        'code': '0062', 'name': '个人贷款净增额', 'category': '保险/贷款',
        'unit': '元', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '活动/任务期间客户零售银行贷款余额的净增加额。反映个人贷款业务增长情况。',
        'biz_use': '个人贷款业务考核',
        'calc_rule': 'CURNT_VAL = SUM(期末LOAN_BAL - 期初LOAN_BAL)\n期初 = 活动/任务开始日\n【v1.3确认】DWS_CUST_ASSE_LIAB.LOAN_BAL已排除普惠/微粒贷/美团贷款，无需额外过滤',
        'source_tables': 'DWS_CUST_ASSE_LIAB (BAL_TYPE=1, LOAN_BAL字段)\n【v1.3】已确认LOAN_BAL不含普惠/微粒贷/美团贷款',
        'stat_cycle': '期初 vs 跑批日',
        'dimensions': '',
        'special_cases': '无需额外过滤逻辑，DWS层LOAN_BAL已排除',
        'confirmed': [
            '[口径] 期末-期初LOAN_BAL',
            '[源表] DWS_CUST_ASSE_LIAB, LOAN_BAL字段',
            '[v1.3确认] LOAN_BAL已排除普惠/微粒贷/美团，无需额外过滤',
        ],
        'pending': [],
        'version': 'v1.3记忆卡片',
    },

    # ---- Critical / Payroll (0063-0064) ----
    {
        'code': '0063', 'name': '临界客户净增', 'category': '临界客户/代发薪',
        'unit': '户', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '上月AUM月日均在临界区间的客户在考核期内的净增数量。用于追踪即将升级的潜力客户经营效果。',
        'biz_use': '临界客户提升经营考核|精准营销转化率',
        'calc_rule': 'CURNT_VAL = COUNT(考核期内,上月AUM月日均在临界区间的客户)\n临界区间定义:\n  临界优质: [4.5万, 5万)\n  临界财富1: [27万, 30万)\n  临界财富2: [45万, 50万)\n  临界贵宾: [90万, 100万)\n  临界私行: [270万, 300万)\n区间规则: 左闭右开 [lower, upper)',
        'source_tables': 'DWS_CUST_ASSE_LIAB (BAL_TYPE=2, AUM月日均)',
        'stat_cycle': '上月AUM月日均为基准，考核期累计净增',
        'dimensions': '',
        'special_cases': '区间边界值使用左闭右开规则(与睡眠户唤醒模块一致)\n若客户上月无AUM数据(AUM IS NULL)，不计入临界客户\n客户从临界区间升级后是否一直保留？仅考核期内新增进入临界区间的客户',
        'confirmed': [
            '[口径] 上月AUM月日均在临界区间',
            '[区间] [4.5,5万)/[27,30万)/[45,50万)/[90,100万)/[270,300万)',
            '[源表] DWS_CUST_ASSE_LIAB, BAL_TYPE=2',
        ],
        'pending': [

        ],
        'version': 'v1.2记忆卡片',
    },
    {
        'code': '0064', 'name': '代发薪客户净增', 'category': '临界客户/代发薪',
        'unit': '户', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '活动/任务期间新签约的代发薪客户数量(按CRM标签中代发口径)。',
        'biz_use': '代发薪业务拓展考核',
        'calc_rule': 'CURNT_VAL = COUNT(考核期内新签代发合同的客户)\nWHERE CTRAKT_TYP=\'代发\'\nAND SIGN_DATE BETWEEN 期初 AND 跑批日',
        'source_tables': 'DWD_CUST_SIGN_CTRAKT (客户签约合同表)\n字段: CTRAKT_TYP, SIGN_DATE',
        'stat_cycle': '期初至跑批日累计',
        'dimensions': '',
        'special_cases': 'CTRAKT_TYP值域待确认(字符串"代发"or编码)',
        'confirmed': [
            '[口径] 考核期内新签代发合同客户',
            '[源表] DWD_CUST_SIGN_CTRAKT',
        ],
        'pending': [
            '[待确认Q4] CTRAKT_TYP值域(字符串or编码?含哪些子类型?)',

        ],
        'version': 'v1.2记忆卡片',
    },

    # ---- Agency income / NPL (0065-0066) ----
    {
        'code': '0065', 'name': '代销业务收入', 'category': '代销收入/不良贷款',
        'unit': '元', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '活动/任务期间个人代销业务产生的全部中间业务收入，包括保险代销佣金、贵金属代销收入、理财代销收入。',
        'biz_use': '代销中间业务收入考核',
        'calc_rule': 'CURNT_VAL = SUM(保险代销收入) + SUM(理财代销收入)\n【v1.3】贵金属代销收入暂不开发，存储过程预留注释和扩展位置\n保险: DWD_ACCT_INSUR中佣金/手续费字段\n理财: DWD_ACCT_FIN中代销手续费',
        'source_tables': 'DWD_ACCT_INSUR (保险) + DWD_ACCT_FIN (理财)\n【v1.3】贵金属暂不开发，待数据源确认后补充',
        'stat_cycle': '期初至跑批日累计',
        'dimensions': '',
        'special_cases': '【v1.3】贵金属代销收入本版本不开发，存储过程注释 TODO: 贵金属代销收入接入',
        'confirmed': [
            '[v1.3口径] 保险代销 + 理财代销（贵金属暂不开发）',
            '[保险] DWD_ACCT_INSUR',
            '[理财] DWD_ACCT_FIN',
            '[v1.3] 贵金属代销待数据源确认后补充',
        ],
        'pending': [],
        'version': 'v1.3记忆卡片',
    },
    {
        'code': '0066', 'name': '个贷新形成不良贷款率', 'category': '代销收入/不良贷款',
        'unit': '%', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '考核期初为正常类(正常+关注)的贷款，在期末变为不良类(次级+可疑+损失)的金额占期初正常贷款总额的比例。',
        'biz_use': '信贷资产质量监控|风险管理考核',
        'calc_rule': 'CURNT_VAL = SUM(期初CATE_5LVL IN (1,2) AND 期末CATE_5LVL IN (3,4,5)的贷款余额)\n          / SUM(期初CATE_5LVL IN (1,2)的贷款总额) * 100\n五级分类: 1=正常, 2=关注, 3=次级, 4=可疑, 5=损失',
        'source_tables': 'DWD_ACCT_LOAN (贷款账户表)\n字段: CATE_5LVL(五级分类), LOAN_BAL(贷款余额)',
        'stat_cycle': '期初分类 vs 期末分类',
        'dimensions': '',
        'special_cases': '期初已为不良(3/4/5)的不计入分子分母\n分母为0时CURNT_VAL=0或NULL',
        'confirmed': [
            '[口径] 新形成不良余额/期初正常总额',
            '[源表] DWD_ACCT_LOAN, CATE_5LVL',
            '[分类] 1正常2关注3次级4可疑5损失',
        ],
        'pending': [

        ],
        'version': 'v1.2记忆卡片',
    },

    # ---- MB / Payment / Card (0067-0072) ----
    {
        'code': '0067', 'name': '手机银行活跃客户数', 'category': '手机银行/支付/绑卡',
        'unit': '户', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '当年登录过手机银行的活跃客户数量。反映手机银行渠道的客户活跃度和粘性。',
        'biz_use': '手机银行渠道活跃度考核|数字化转型指标',
        'calc_rule': 'CURNT_VAL = COUNT(DISTINCT 当年有登录记录的客户)\n登录记录: ODS mbk_cust_log_login\n过滤: lgn_status=\'1\'(成功)\n活跃阈值: 待确认(至少1次？N次？)',
        'source_tables': 'ODS: mbk_cust_log_login (手机银行登录日志)\n字段: cust_no, lgn_date, lgn_status',
        'stat_cycle': '当年1月1日至跑批日',
        'dimensions': '',
        'special_cases': 'cust_no(电子银行客户号)需映射到CRM CUST_ID\n活跃频次阈值待确认',
        'confirmed': [
            '[口径] 当年登录过的客户数',
            '[源表] ODS mbk_cust_log_login',
            '[过滤] lgn_status=\'1\'(成功)',
        ],
        'pending': [
            '[待确认Q3] 活跃是否有最低登录次数/天数阈值？',
            '[待确认] cust_no->CUST_ID映射方式',

        ],
        'version': 'v1.2记忆卡片',
    },
    {
        'code': '0068', 'name': '支付结算年度价值商户数', 'category': '手机银行/支付/绑卡',
        'unit': '户', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '同时满足客户资产指标和商户自身交易量指标的价值商户数量。',
        'biz_use': '支付结算商户质量考核',
        'calc_rule': '价值商户判定 = 客户资产达标 AND 商户交易量达标\n客户资产: 小微商户(AUM>=2万or活期>=1万) /个体商户(AUM>=5万or活期>=2.5万)\n交易量: 近30天条码收单交易笔数>=10 AND 金额>=100元\n商户类型过滤: uepp_pay_mct_info.mct_type IN (\'personage\',\'smallBusinesses\')',
        'source_tables': 'DWS_CUST_ASSE_LIAB(客户AUM/活期)\nODS uepp_pay_order_info(交易流水)\nODS uepp_pay_mct_info(商户类型: mct_type=personage/smallBusinesses)',
        'stat_cycle': '客户资产: 当月月日均 | 交易量: 近30天(含当日)',
        'dimensions': '',
        'special_cases': '商户分小微/个体两个门槛，按商户类型分别判断\n商户与个人客户的关联逻辑待确认',
        'confirmed': [
            '[口径] 资产+交易量双达标',
            '[资产] 小微:2万/1万, 个体:5万/2.5万',
            '[交易] 近30天>=10笔AND>=100元',
            '[源表] uepp_pay_mct_info.mct_type区分商户类型',
            '[Q1已解决] mct_type: personage=个体商户, smallBusinesses=小微商户',
        ],
        'pending': [
            '[待确认] 商户->CRM客户号映射方式',
            '[待确认] AUM/活期指关联个人客户AUM还是商户结算存款？',

        ],
        'version': 'v1.3.1记忆卡片',
    },
    {
        'code': '0069', 'name': '二维码收款个人活跃商户结算存款留存率', 'category': '手机银行/支付/绑卡',
        'unit': '%', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '年累计交易金额>=500元的商户，存款日均余额占年累计交易金额的比例。衡量商户资金在行内的留存程度。',
        'biz_use': '商户资金留存率考核',
        'calc_rule': 'CURNT_VAL = 年日均存款 / 年累计交易金额 * 100\n筛选: 年累计交易金额 >= 500元\n商户类型: mct_type IN (\'personage\',\'smallBusinesses\')\n年日均存款: 商户关联存款账户的当年日均余额',
        'source_tables': 'ODS uepp_pay_order_info(交易金额) + DWD_ACCT_DEPO(存款) + uepp_pay_mct_info(商户类型)',
        'stat_cycle': '当年累计',
        'dimensions': '',
        'special_cases': '分母年累计交易金额为0时结果为NULL',
        'confirmed': [
            '[口径] 年日均存款/年累计交易量',
            '[筛选] 年累计交易>=500元',
            '[源表] uepp+ACCT_DEPO',
            '[Q1已解决] 仅统计个体工商户+小微商户: mct_type IN (\'personage\',\'smallBusinesses\')',
        ],
        'pending': [
            '[待确认] 同0068的商户关联问题',

        ],
        'version': 'v1.3.1记忆卡片',
    },
    {
        'code': '0070', 'name': '银行卡三方支付绑卡', 'category': '手机银行/支付/绑卡',
        'unit': '张', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '我行银行卡首次绑定微信支付、支付宝、云闪付等三方支付的数量(按发卡机构计算)。',
        'biz_use': '三方支付绑卡推广考核',
        'calc_rule': 'CURNT_VAL = COUNT(考核期内首次绑定的卡数)\n"首次"定义待确认(该卡首次/该客户首次?)\n平台范围: 微信+支付宝+云闪付',
        'source_tables': 'ODS: mbk_cust_acct (手机银行账户绑定)\n字段: cust_no, acct, acct_add_date, acct_add_chnl',
        'stat_cycle': '活动/任务期间',
        'dimensions': '',
        'special_cases': '第一次"首次"判定逻辑是关键(见Q2)\ncust_no需映射到CUST_ID\n按发卡机构聚合',
        'confirmed': [
            '[口径] 首次绑定三方支付的卡数',
            '[源表] ODS mbk_cust_acct',
            '[平台] 微信+支付宝+云闪付',
        ],
        'pending': [
            '[待确认Q2-阻塞] "首次"判定: 卡级首次?客户级首次?有别的绑卡表?',
            '[待确认Q2] mbk_cust_acct如何区分微信/支付宝/云闪付(哪个字段)?',
            '[待确认Q2] cust_no->CUST_ID映射方式',

        ],
        'version': 'v1.2记忆卡片',
    },
    {
        'code': '0071', 'name': '银行卡三方支付绑卡率', 'category': '手机银行/支付/绑卡',
        'unit': '%', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '本年新增持卡且年龄<=70岁的客户中，已绑定三方支付的比例。',
        'biz_use': '绑卡推广渗透率考核',
        'calc_rule': 'CURNT_VAL = 绑卡数(0070) / 本年新增持卡客户数(<=70岁) * 100\n本年新增持卡: 开户日期在当年的客户\n年龄: DWD_CUST_INDV_INFO中计算,<=70岁',
        'source_tables': 'mbk_cust_acct(绑卡) + DWD_CUST_INDV_INFO(年龄/BIRTH_DATE)',
        'stat_cycle': '当年累计',
        'dimensions': '',
        'special_cases': '分母为0时结果为NULL\n年龄以跑批日为基准计算',
        'confirmed': [
            '[口径] 绑卡数/本年新增持卡客户(<=70岁)',
            '[年龄] DWD_CUST_INDV_INFO.BIRTH_DATE, <=70岁',
        ],
        'pending': [
            '[待确认Q2] "本年新增持卡"指开户日期在当年?还是首次绑卡日期在当年?',
            '[待确认Q2] 同0070的其他待确认项',

        ],
        'version': 'v1.2记忆卡片',
    },
    {
        'code': '0072', 'name': '活跃卡数', 'category': '手机银行/支付/绑卡',
        'unit': '张', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '近180天内该卡号所属根卡下任意卡片有发生消费、取现、还款等主动动账交易的有效卡片数量。',
        'biz_use': '银行卡活跃度考核',
        'calc_rule': 'CURNT_VAL = COUNT(DISTINCT 近180天有主动动账的卡号)\n主动动账: DWD_TX_ASET.JIOYCFFS=\'0\'\n窗口: TX_DATE BETWEEN T-179 AND T',
        'source_tables': 'DWD_TX_ASET (交易流水表)\n字段: JIOYCFFS, TX_DATE, 卡号',
        'stat_cycle': '近180天(含当日)',
        'dimensions': '',
        'special_cases': '根卡概念: 卡号所属根卡下任意卡片有交易即计入\n同一根卡下多张子卡不重复计数',
        'confirmed': [
            '[口径] 近180天有主动动账的卡数',
            '[源表] DWD_TX_ASET, JIOYCFFS=\'0\'',
            '[窗口] 180天',
        ],
        'pending': [
            '[待确认] 根卡与子卡的关联逻辑(哪个表哪些字段)',

        ],
        'version': 'v1.2记忆卡片',
    },

    # ---- Report platform (0073-0079) ----
    {
        'code': '0073', 'name': '手机银行客户数新增个人客户数', 'category': '综合报表平台',
        'unit': '户', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '综合报表平台提供的手机银行新增客户数。(本期不开发)',
        'biz_use': '待外部接口',
        'calc_rule': '外部数据，本期不处理',
        'source_tables': '综合报表平台(外部)',
        'stat_cycle': '不适用',
        'dimensions': '不适用',
        'special_cases': '本期不处理，待综合报表平台接口就绪',
        'confirmed': ['[状态] 本期不开发'],
        'pending': ['综合报表平台数据接口时间待定'],
        'version': 'v1.1记忆卡片(本期不处理)',
    },
    {
        'code': '0074', 'name': '手机银行交易金额', 'category': '综合报表平台',
        'unit': '元', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '综合报表平台提供的手机银行交易金额。(本期不开发)',
        'biz_use': '待外部接口',
        'calc_rule': '外部数据，本期不处理',
        'source_tables': '综合报表平台(外部)',
        'stat_cycle': '不适用',
        'dimensions': '不适用',
        'special_cases': '本期不处理',
        'confirmed': ['[状态] 本期不开发'],
        'pending': ['综合报表平台数据接口时间待定'],
        'version': 'v1.1记忆卡片(本期不处理)',
    },
    {
        'code': '0075', 'name': '手机银行交易笔数', 'category': '综合报表平台',
        'unit': '笔', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '综合报表平台提供的手机银行交易笔数。(本期不开发)',
        'biz_use': '待外部接口',
        'calc_rule': '外部数据，本期不处理',
        'source_tables': '综合报表平台(外部)',
        'stat_cycle': '不适用',
        'dimensions': '不适用',
        'special_cases': '本期不处理',
        'confirmed': ['[状态] 本期不开发'],
        'pending': ['综合报表平台数据接口时间待定'],
        'version': 'v1.1记忆卡片(本期不处理)',
    },
    {
        'code': '0076', 'name': '一码付收款户数', 'category': '综合报表平台',
        'unit': '户', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '综合报表平台提供的一码付收款客户数。(本期不开发)',
        'biz_use': '待外部接口',
        'calc_rule': '外部数据，本期不处理',
        'source_tables': '综合报表平台(外部)',
        'stat_cycle': '不适用',
        'dimensions': '不适用',
        'special_cases': '本期不处理',
        'confirmed': ['[状态] 本期不开发'],
        'pending': ['综合报表平台数据接口时间待定'],
        'version': 'v1.1记忆卡片(本期不处理)',
    },
    {
        'code': '0077', 'name': '一码付收款新增商户数', 'category': '综合报表平台',
        'unit': '户', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '综合报表平台提供的一码付新增商户数。(本期不开发)',
        'biz_use': '待外部接口',
        'calc_rule': '外部数据，本期不处理',
        'source_tables': '综合报表平台(外部)',
        'stat_cycle': '不适用',
        'dimensions': '不适用',
        'special_cases': '本期不处理',
        'confirmed': ['[状态] 本期不开发'],
        'pending': ['综合报表平台数据接口时间待定'],
        'version': 'v1.1记忆卡片(本期不处理)',
    },
    {
        'code': '0078', 'name': '一码付收款交易金额', 'category': '综合报表平台',
        'unit': '元', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '综合报表平台提供的一码付交易金额。(本期不开发)',
        'biz_use': '待外部接口',
        'calc_rule': '外部数据，本期不处理',
        'source_tables': '综合报表平台(外部)',
        'stat_cycle': '不适用',
        'dimensions': '不适用',
        'special_cases': '本期不处理',
        'confirmed': ['[状态] 本期不开发'],
        'pending': ['综合报表平台数据接口时间待定'],
        'version': 'v1.1记忆卡片(本期不处理)',
    },
    {
        'code': '0079', 'name': '一码付收款交易笔数', 'category': '综合报表平台',
        'unit': '笔', 'freq': '日', 'object': '理财经理/机构', 'dim': '营销活动/目标任务',
        'biz_meaning': '综合报表平台提供的一码付交易笔数。(本期不开发)',
        'biz_use': '待外部接口',
        'calc_rule': '外部数据，本期不处理',
        'source_tables': '综合报表平台(外部)',
        'stat_cycle': '不适用',
        'dimensions': '不适用',
        'special_cases': '本期不处理',
        'confirmed': ['[状态] 本期不开发'],
        'pending': ['综合报表平台数据接口时间待定'],
        'version': 'v1.1记忆卡片(本期不处理)',
    },
]

# ================================================================
# Sheet 1: Full specification
# ================================================================
ws1 = wb.active
ws1.title = '指标技术规格'

ws1.merge_cells('A1:M1')
style_cell(ws1, 1, 1, '指标数据统计(ADS_STAT_INDX_DATA) - 技术规格与确认项', Font(name='Microsoft YaHei', bold=True, size=14, color='1F4E79'))
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 30

ws1.merge_cells('A2:M2')
style_cell(ws1, 2, 1, '版本: v1.3 | 日期: 2026-08-04 | 27个开发指标 + 7个本期不处理 (v1.3口径: 储蓄=AUM, 贵金属暂不开发, LOAN已排除)', Font(name='Microsoft YaHei', size=9, color='666666'))
ws1['A2'].alignment = Alignment(horizontal='center')

h1 = ['编号', '指标名称', '类别', '单位', '统计对象', '统计维度', '业务理解', '技术口径(计算规则)', '数据来源', '统计周期', '维度划分', '已确认项', '待确认项']
for c, h in enumerate(h1, 1):
    ws1.cell(row=4, column=c, value=h)
style_hdr(ws1, 4, len(h1))
ws1.row_dimensions[4].height = 22

row = 5
for idx in indicators:
    vals = [
        idx['code'], idx['name'], idx['category'], idx['unit'],
        idx['object'], idx['dim'],
        f'{idx["biz_meaning"]}\n\n【用途】{idx["biz_use"]}',
        idx['calc_rule'],
        idx['source_tables'],
        idx['stat_cycle'],
        '见"维度划分详情"Sheet' if not idx['dimensions'] else idx['dimensions'],
        '\n'.join(idx['confirmed']),
        '\n'.join(idx['pending']) if idx['pending'] else '无',
    ]
    for c, v in enumerate(vals, 1):
        is_alt = (row % 2 == 1)
        f = normal_font
        fl = alt_fill if is_alt else None
        style_cell(ws1, row, c, v, f, fl)
    # Color code the confirmed/pending columns
    ws1.cell(row=row, column=12).fill = confirmed_fill
    if idx['pending']:
        ws1.cell(row=row, column=13).fill = pending_fill
    ws1.row_dimensions[row].height = max(120, 20 * len(idx['calc_rule'].split('\n')))
    row += 1

# Handle categories - also add at end for reference
categories = ['储蓄存款', '客户层级', '理财产品', '代销理财', '保险/贷款', '临界客户/代发薪', '代销收入/不良贷款', '手机银行/支付/绑卡', '综合报表平台']

set_widths(ws1, [8, 28, 12, 6, 14, 16, 42, 48, 36, 20, 30, 40, 40])

# ================================================================
# Sheet 2: By category
# ================================================================
ws2 = wb.create_sheet('按类别分组')

ws2.merge_cells('A1:J1')
style_cell(ws2, 1, 1, '按指标类别分组的业务理解与技术口径', Font(name='Microsoft YaHei', bold=True, size=13, color='1F4E79'))
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 26

h2 = ['编号', '指标名称', '单位', '业务理解', '计算规则', '关键字段/表', '已确认', '待确认', '风险', '状态']
row = 3
for cat in categories:
    cat_items = [i for i in indicators if i['category'] == cat]
    if not cat_items:
        continue
    section_row(ws2, row, len(h2), f'■ {cat} ({len(cat_items)}个指标)', cat_fill)
    row += 1
    for c, h in enumerate(h2, 1):
        style_cell(ws2, row, c, h, hdr_font, hdr_fill, wrap_center)
    ws2.row_dimensions[row].height = 22
    row += 1

    for idx in cat_items:
        pending_count = len(idx['pending'])
        if idx['code'] in ('0073','0074','0075','0076','0077','0078','0079'):
            status = '本期不处理'
            risk = '-'
        elif pending_count == 0:
            status = '可开发'
            risk = '低'
        elif any('阻塞' in p for p in idx['pending']):
            status = '阻塞'
            risk = '高'
        else:
            status = '待确认'
            risk = '中' if pending_count > 1 else '低'

        vals = [
            idx['code'], idx['name'], idx['unit'],
            idx['biz_meaning'][:120] + '...' if len(idx['biz_meaning']) > 120 else idx['biz_meaning'],
            idx['calc_rule'][:150] + '...' if len(idx['calc_rule']) > 150 else idx['calc_rule'],
            idx['source_tables'][:100] + '...' if len(idx['source_tables']) > 100 else idx['source_tables'],
            f'{len(idx["confirmed"])}项',
            f'{len(idx["pending"])}项' if idx['pending'] else '0项',
            risk, status,
        ]
        for c, v in enumerate(vals, 1):
            is_alt = ((row - 5) % 2 == 1)
            fl = alt_fill if is_alt else None
            style_cell(ws2, row, c, v, normal_font, fl, wrap_center if c >= 6 else wrap_top)
        if risk == '高':
            ws2.cell(row=row, column=9).font = red_font
        if status == '可开发':
            ws2.cell(row=row, column=10).font = green_font
        elif status == '阻塞':
            ws2.cell(row=row, column=10).font = red_font
        ws2.row_dimensions[row].height = 75
        row += 1
    row += 1

set_widths(ws2, [8, 28, 6, 45, 50, 38, 10, 10, 8, 10])

# ================================================================
# Sheet 3: Pending items summary
# ================================================================
ws3 = wb.create_sheet('待确认项汇总')

ws3.merge_cells('A1:F1')
style_cell(ws3, 1, 1, '待确认项按问题编号汇总', Font(name='Microsoft YaHei', bold=True, size=13, color='1F4E79'))
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')

h3 = ['问题编号', '问题', '影响指标', '阻塞等级', '涉及指标数', '确认对象']
row = 3
for c, h in enumerate(h3, 1):
    style_cell(ws3, row, c, h, hdr_font, hdr_fill, wrap_center)
ws3.row_dimensions[row].height = 22
row += 1

pending_summary = [
    ['Q1', '商户小微/个体区分标识及商户->客户映射', '0068, 0069', '已解决', 2, 'mct_type确认: personage=个体商户, smallBusinesses=小微商户; 商户映射待确认'],
    ['Q2', '绑卡"首次绑定"判断逻辑及平台区分', '0070, 0071', 'P0', 2, '手机银行系统负责人'],
    ['Q3', '手机银行登录"活跃"频次阈值', '0067', 'P1', 1, '手机银行业务负责人'],
    ['Q4', '代发薪签约类型CTRAKT_TYP值域', '0064', 'P1', 1, 'CRM数据仓库负责人'],
    ['Q7', '贵金属代销收入源表及字段', '0065', 'P0', 1, '贵金属业务负责人'],
    ['—', 'DWS层LOAN_BAL是否已排除普惠/微粒贷/美团贷款', '0062', 'P1', 1, 'CRM数据仓库负责人'],
    ['—', '根卡与子卡的关联逻辑(表/字段)', '0072', 'P1', 1, 'CRM数据仓库负责人'],
    ['—', 'mbk_cust_no->CRM_CUST_ID映射方式', '0067,0070,0071', 'P1', 3, 'CRM数据仓库负责人'],
    ['✓ Q5', '借贷管户(MNG_TYP=2)是否纳入目标任务', '全部路径B', '已解决', 27, '不含MNG_TYP=2,统一MNG_TYP=1'],
    ['✓ Q6', 'INDX_ID存储格式', '全部路径A', '已解决', 27, 'INDX_0046格式(与INDX_CODE一致)'],
    ['✓ Q8', 'RSV_OBJ=机构时无效客户过滤', '全部路径B', '已解决', 27, '不需要过滤,各源表天然排除'],
]

for item in pending_summary:
    for c, v in enumerate(item, 1):
        style_cell(ws3, row, c, v, normal_font, alt_fill if row % 2 == 1 else None, wrap_center if c >= 2 else wrap_top)
    if item[3] == 'P0':
        ws3.cell(row=row, column=4).font = red_font
    elif item[3] == '已解决':
        ws3.cell(row=row, column=4).font = green_font
    ws3.row_dimensions[row].height = 35
    row += 1

set_widths(ws3, [10, 48, 28, 10, 14, 26])

# ================================================================
# Sheet 4: Data source mapping
# ================================================================
ws4 = wb.create_sheet('数据源映射')

ws4.merge_cells('A1:G1')
style_cell(ws4, 1, 1, '指标 -> 源表/字段映射', Font(name='Microsoft YaHei', bold=True, size=13, color='1F4E79'))
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')

h4 = ['表名', '所属层', '关联指标', '关键字段', '用途', 'DDL状态', '备注']
row = 3
for c, h in enumerate(h4, 1):
    style_cell(ws4, row, c, h, hdr_font, hdr_fill, wrap_center)
ws4.row_dimensions[row].height = 22
row += 1

sources = [
    ['DWS_CUST_ASSE_LIAB', 'DWS', '0046-0051,0055-0059,0062,0063,0068',
     'PERSN_LEGAL_BK_CODE,CUST_ID,ORG_ID,DEPO_CURNT_DEPO_BAL,FIXD_DEPO_BAL,FIN_BAL,INSUR_BAL,LOAN_BAL,AUM_BAL,LEHUI_BAL,LARGEDP_BAL',
     '储蓄/理财/保险/贷款/AUM余额;BAL_TYPE=1(日)/2(月日均)/4(年日均)', '已存在', ''],
    ['DWS_CUST_LVL_INFO', 'DWS', '0052-0054',
     'CUST_ID,PERSN_LEGAL_BK_CODE,CUST_LVL',
     '客户等级(0-7级)', '已存在', ''],
    ['DWD_ACCT_FIN', 'DWD', '0055-0060,0065',
     'PERSN_LEGAL_BK_CODE,CUST_ID,FIN_AMT,PRDKT_CATE_BIG,ISSU_DATE',
     '理财产品余额/购买/赎回/代销收入', '已存在', ''],
    ['DWD_ACCT_INSUR', 'DWD', '0061,0065',
     'PERSN_LEGAL_BK_CODE,CUST_ID,NEW_INSUR_AMT,CANCL_INSUR_DATE,LAST_TX_DATE',
     '保险首期保费/退保/代销收入', '已存在', 'v2.0'],
    ['DWD_ACCT_LOAN', 'DWD', '0062,0066',
     'PERSN_LEGAL_BK_CODE,CUST_ID,LOAN_BAL,CATE_5LVL',
     '贷款余额/五级分类', '已存在', '排除范围待确认'],
    ['DWD_CUST_SIGN_CTRAKT', 'DWD', '0064',
     'CUST_ID,CTRAKT_TYP,SIGN_DATE',
     '代发薪签约合同', '已存在', 'CTRAKT_TYP值域待确认'],
    ['DWD_CUST_MAN', 'DWD', '全部路径B',
     'CUST_ID,MNGR_POST_ID,MNG_TYP,PERSN_LEGAL_BK_CODE',
     '客户-经理管户(MNG_TYP=1/2)', '已存在', ''],
    ['DWD_SYS_ORG', 'DWD', '全部路径A+B',
     'ORG_ID,SUP_ORG_ID,ORG_NAME',
     '机构层级(递归汇总)', '已存在', ''],
    ['DWD_SYS_POST', 'DWD', '全部路径A+B',
     'POST_ID,ORG_ID,JOB_CLS',
     '职位-机构关联', '已存在', ''],
    ['DWD_TX_ASET', 'DWD', '0072',
     'PERSN_LEGAL_BK_CODE,CUST_ID,JIOYCFFS,TX_DATE',
     '交易流水(主动动账JIOYCFFS=0)', '已存在', ''],
    ['DWD_CUST_INDV_INFO', 'DWD', '0071',
     'CUST_ID,BIRTH_DATE,OPEN_DATE',
     '客户年龄(<=70岁)及开户日期', '已存在', ''],
    ['DWD_ACCT_DEPO', 'DWD', '0069',
     'PERSN_LEGAL_BK_CODE,CUST_ID,INTRI_BGN_DATE',
     '存款账户(商户存款留存率)', '已存在', ''],
    ['DWD_MKT_ACT_INFO', 'DWD', '全部路径A',
     'MKT_ACT_ID,ACT_TYP,ACT_BGN_DATE,STATIS_STOP_DATE',
     '营销活动(ACT_TYP=综合方案/价值提升)', '已存在', ''],
    ['DWD_MKT_ACT_ORG_REL', 'DWD', '全部路径A',
     'MKT_ACT_ID,PRTSPT_ORG',
     '活动-机构关联', '已存在', ''],
    ['DWD_MKT_TSK_INFO', 'DWD', '全部路径A,0047',
     'MKT_ACT_ID,CUST_ID,MKT_PERSN,MKT_PERSN_ORG,BASE_VAL',
     '营销任务(客户+基数)', '已存在', ''],
    ['DWD_MKT_INDX_TSK', 'DWD', '全部路径B',
     'TSK_ID,RSV_OBJ,RSV_OBJ_ID,TSK_BGN_DATE,CREAT_ORG',
     '指标任务(接收对象)', '已存在', ''],
    ['DWD_MKT_ACT_TARGT', 'DWD', '全部路径A',
     'MKT_ACT_ID,INDX_ID',
     '活动-指标关联(新增)', '已确认', 'INDX_0046格式'],
    ['mbk_cust_log_login', 'ODS', '0067',
     'cust_no,lgn_date,lgn_status,lgn_chnl',
     '手机银行登录日志', '已存在', '活跃阈值待确认Q3'],
    ['mbk_cust_acct', 'ODS', '0070,0071',
     'cust_no,acct,acct_add_date,acct_add_chnl',
     '手机银行绑卡记录', '已存在', '首次绑定逻辑待确认Q2'],
    ['uepp_pay_order_info', 'ODS', '0068,0069',
     'order_id,mct_id,order_amt,pay_time,status,isscode,consumer_id',
     '支付交易流水', '已存在', ''],
    ['uepp_pay_mct_info', 'ODS', '0068,0069',
     'mct_id,mct_type,name,org_id,job_id,sign_date',
     '商户类型枚举: company=企业,institution=党政机关/事业单位,otherOrganizations=其他组织,personage=个体商户,smallBusinesses=小微商户。0068/0069过滤: mct_type IN (\'personage\',\'smallBusinesses\')', '已确认v1.3.1', ''],
    ['贵金属代销表', '待定', '0065',
     '待确认',
     '贵金属代销交易+收入', '缺失', '待确认Q7'],
]

for item in sources:
    for c, v in enumerate(item, 1):
        is_alt = (row % 2 == 1)
        fl = alt_fill if is_alt else None
        f = red_font if item[5] == '缺失' else (orange_font if '待确认' in item[5] else normal_font)
        style_cell(ws4, row, c, v, f, fl, wrap_top)
    ws4.row_dimensions[row].height = 40
    row += 1

set_widths(ws4, [26, 8, 24, 60, 38, 14, 30])

# ================================================================
# Sheet 5: Dimension details
# ================================================================
ws5 = wb.create_sheet('维度划分详情')

ws5.merge_cells('A1:B1')
style_cell(ws5, 1, 1, '指标维度划分详情 — 表关联关系与字段来源', Font(name='Microsoft YaHei', bold=True, size=14, color='1F4E79'))
ws5['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[1].height = 30

ws5.merge_cells('A2:B2')
style_cell(ws5, 2, 1, '版本: v1.3 | 日期: 2026-08-04 | 适用于 INDX_0046-0072 (27个开发指标) | 路径重构+MNG_TYP=1', Font(name='Microsoft YaHei', size=9, color='666666'))
ws5['A2'].alignment = Alignment(horizontal='center')

h5 = ['维度类别', '详情（表名.字段名）']
for c, h in enumerate(h5, 1):
    style_cell(ws5, 4, c, h, hdr_font, hdr_fill, wrap_center)
ws5.row_dimensions[4].height = 22

dim_rows = [
    ('一、统计口径维度 (STATIS_CALIB)', '', True),
    ('字段值', 'ADS_STAT_INDX_DATA.STATIS_CALIB = "营销活动" / "目标任务"', False),
    ('数据路径A: 营销活动', '入参表: DWD_MKT_ACT_INFO(MKT_ACT_ID, ACT_TYP, ACT_BGN_DATE, STATIS_STOP_DATE)\n入参表: DWD_MKT_TSK_INFO(MKT_ACT_ID, CUST_ID, MKT_PERSN, MKT_PERSN_ORG, BASE_VAL)\n入参表: DWD_MKT_ACT_TARGT(MKT_ACT_ID, INDX_ID=INDX_0046格式)\n过滤: ACT_TYP IN ("综合方案","价值提升")，ACT_TYP="客户维系"不产出\n【v1.3】机构来源: MKT_PERSN_ORG（客户经理所在机构），非PRTSPT_ORG', False),
    ('数据路径B: 目标任务', '入参表: DWD_MKT_INDX_TSK(TSK_ID, MKT_ACT_ID, RSV_OBJ, RSV_OBJ_ID, TSK_BGN_DATE, CREAT_ORG)\nRSV_OBJ=机构(0): RSV_OBJ_ID → DWD_SYS_ORG.ORG_ID → DWS_CUST_LVL_INFO.ORG_ID → 全量客户 ← 【v1.3】\nRSV_OBJ=理财经理(1): RSV_OBJ_ID → DWD_CUST_MAN(MNG_TYP=1, MNGR_POST_ID) → 该经理名下客户', False),

    ('二、数据归属维度 (DATA_BLNG)', '', True),
    ('字段值', 'ADS_STAT_INDX_DATA.DATA_BLNG = 机构编号 / 客户经理编号(POST_ID)', False),
    ('机构行产出规则', '机构编号 = DWD_SYS_ORG.ORG_ID\n包含范围: 该机构SUP_ORG_ID递归下级所有机构 ↓\n机构→客户: DWD_CUST_MAN(MNG_TYP=1) → 客户号', False),
    ('客户经理行产出规则', 'POST_ID = DWD_SYS_POST.POST_ID\n取数路径: DWD_CUST_MAN.MNGR_POST_ID = POST_ID AND MNG_TYP=1', False),

    ('四、统计对象维度 (STATIS_DIM)', '', True),
    ('路径A值', 'ADS_STAT_INDX_DATA.STATIS_DIM = MKT_ACT_ID(DWD_MKT_ACT_INFO)', False),
    ('路径B值', 'ADS_STAT_INDX_DATA.STATIS_DIM = TSK_ID(DWD_MKT_INDX_TSK)', False),

    ('五、时间维度', '', True),
    ('DATA_DATE', 'ADS_STAT_INDX_DATA.DATA_DATE = V_SYSDAT(VARCHAR(8), YYYYMMDD, 跑批日期)', False),
    ('期初基准(路径A)', 'DWD_MKT_ACT_INFO.ACT_BGN_DATE → 活动开始日', False),
    ('期初基准(路径B)', 'DWD_MKT_INDX_TSK.TSK_BGN_DATE → 任务开始日', False),
    ('期末', 'DATA_DATE = V_SYSDAT (跑批当日)', False),
    ('年初基准(0046)', 'DWS_CUST_ASSE_LIAB.DATA_DATE = 当年1月1日', False),
    ('上月基准(0048)', 'DWS_CUST_ASSE_LIAB.DATA_DATE = 上月最后一天', False),
    ('上季基准(0049)', 'DWS_CUST_ASSE_LIAB.DATA_DATE = 上季最后一天', False),
    ('统计周期', '[ACT_BGN_DATE/TSK_BGN_DATE, STATIS_STOP_DATE/DATA_DATE] 每日跑批', False),

    ('五、客户维度', '', True),
    ('客户号(CUST_ID)', '通用字段，贯穿所有源表:\n  DWS_CUST_ASSE_LIAB.CUST_ID\n  DWD_CUST_MAN.CUST_ID\n  DWD_CUST_LVL_INFO.CUST_ID\n  DWD_CUST_INDV_INFO.CUST_ID\n  DWD_ACCT_FIN.CUST_ID\n  DWD_ACCT_INSUR.CUST_ID\n  DWD_ACCT_LOAN.CUST_ID\n  DWD_CUST_SIGN_CTRAKT.CUST_ID\n  DWD_MKT_TSK_INFO.CUST_ID\n【v1.3】mbk→CRM映射: mbk_cust_info.cust_core_no = CUST_ID (核心客户号)\n          mbk_cust_info.cust_no = 手机银行客户号 (用于JOIN mbk_cust_log_login/acct)', False),
    ('法人行号(PERSN_LEGAL_BK_CODE)', '所有产出行均携带，取数优先级:\n  1. 优先从DWD_CUST_MAN.PERSN_LEGAL_BK_CODE\n  2. 各源表自带的PERSN_LEGAL_BK_CODE', False),
    ('客户等级(CUST_LVL)', 'DWD_CUST_LVL_INFO: 0=睡眠户/1=大众/2=潜力/3=优质/4=财富1/5=财富2/6=贵宾/7=私行\n判断基准: AUM月日均(非当前余额)\n指标0052: 期初LVL<4→期末>=4\n指标0053: 期初LVL<6→期末>=6\n指标0054: 期初LVL<7→期末=7', False),
    ('客户年龄(绑卡率0071)', 'DWD_CUST_INDV_INFO.BIRTH_DATE → 计算年龄(以跑批日为基准)\n过滤: 年龄 <= 70岁', False),

    ('六、机构维度 (递归向上汇总)', '', True),
    ('机构层级表', 'DWD_SYS_ORG(ORG_ID, SUP_ORG_ID, ORG_NAME)\n递归规则: CONNECT BY PRIOR SUP_ORG_ID = ORG_ID\n从叶子机构(网点)向上递归到根(总行)', False),
    ('路径A机构取数', 'DWD_MKT_TSK_INFO.MKT_PERSN_ORG → 对应DWD_SYS_ORG.ORG_ID（叶子机构）\n汇总: 每个MKT_PERSN_ORG叶子机构+其递归上级机构分别产出统计行 ← 【v1.3: 从MKT_PERSN_ORG取，不再从PRTSPT_ORG】', False),
    ('路径B机构取数', 'RSV_OBJ=机构(0) → RSV_OBJ_ID → DWD_SYS_ORG.ORG_ID\n客户获取: DWS_CUST_LVL_INFO.ORG_ID = RSV_OBJ_ID → 该机构下所有客户 ← 【v1.3: 直接从客户等级表取，不通过管户关系】', False),
    ('机构汇总方式', '上级机构 = 所有下级机构叶子行 + 所有下级机构下客户经理行的聚合', False),

    ('七、客户经理维度', '', True),
    ('职位表', 'DWD_SYS_POST(POST_ID, ORG_ID, JOB_CLS, POST_STATE)\nJOB_CLS="C"(客户经理岗位)\nPOST_STATE=有效状态', False),
    ('客户管户表', 'DWD_CUST_MAN(CUST_ID, MNGR_POST_ID, MNG_TYP, PERSN_LEGAL_BK_CODE)\nMNG_TYP=1(理财管户) — 统一使用(v1.2 Q5决议)', False),
    ('管户关联规则', 'DWD_CUST_MAN.MNGR_POST_ID = DWD_SYS_POST.POST_ID\nAND DWD_CUST_MAN.MNG_TYP = 1', False),
    ('路径A经理取数', 'DWD_MKT_TSK_INFO.MKT_PERSN → DWD_SYS_POST.POST_ID → DWD_SYS_POST.ORG_ID\n产出: 经理行(DATA_BLNG=MKT_PERSN) + 所属机构行(DATA_BLNG=MKT_PERSN_ORG递归向上)\n说明: 经理行是该经理名下客户汇总；机构行是叶子机构→上级机构汇总 ← 【v1.3】', False),
    ('为何客户经理维度会产出机构行?', '双纬度考核模式:\n1. 客户经理行: 考核个人业绩 — 该经理名下所有客户指标汇总\n2. 机构行: 考核团队/机构整体 — 叶子机构及上级各层汇总\n上级机构既需要看到每位经理的单人表现，也需要看到一个网点/支行的整体经营情况。\n所以MKT_PERSN_ORG被同时用于产出经理关联的机构行和数据归属。', False),
    ('路径B经理取数', 'DWD_MKT_INDX_TSK.RSV_OBJ=理财经理(1) → RSV_OBJ_ID=经理编号\n产出: RSV_OBJ_ID(接收对象行) + CREAT_ORG(下发机构行)', False),

    ('八、指标维度 (路径A特有)', '', True),
    ('活动-指标关联', 'DWD_MKT_ACT_TARGT(MKT_ACT_ID, INDX_ID)\nINDX_ID格式: INDX_0046 (带INDX_前缀+下划线)\n与ADS_STAT_INDX_DATA.INDX_CODE格式一致(v1.2 Q6决议)', False),
    ('活动类型过滤', 'DWD_MKT_ACT_INFO.ACT_TYP IN ("综合方案","价值提升")\n排除: ACT_TYP="客户维系" (不产出指标)', False),

    ('九、各指标专项维度', '', True),
    ('商户维度(0068/0069)', 'ODS uepp_pay_mct_info.mct_type 完整枚举:\n  company=企业, institution=党政机关/事业单位, otherOrganizations=其他组织,\n  personage=个体商户, smallBusinesses=小微商户\n[Q1已解决] 个体商户+小微商户判定: mct_type IN (\'personage\',\'smallBusinesses\')\nODS uepp_pay_order_info (交易流水: order_amt, pay_time, isscode)\n商户→客户映射: 待确认', False),
    ('产品维度(0057/0060)', 'DWD_ACCT_FIN.PRDKT_CATE_BIG: 产品大类编码（来源: 实际数据分布确认）\n  代码1,2 = 代销理财产品 → 0058/0059/0060 过滤: IN ("1","2")\n  代码3,4 = 自营理财产品 → 归入0055/0056/0057总量\n  0055/0056/0057(总量)不用过滤，涵盖所有编码\n  注意: 08_移动端客户.md的"1,3=开放式"为不同模块定义，不适用', False),
    ('卡维度(0070-0072)', 'ODS mbk_cust_acct.acct(卡号), ODS mbk_cust_acct.acct_add_date(绑定日期)\nDWD_TX_ASET: 交易卡号, JIOYCFFS=\'0\'(主动动账)', False),
    ('保险维度(0061)', 'DWD_ACCT_INSUR.CANCL_INSUR_DATE IS NULL(未退保)\nDWD_ACCT_INSUR.LAST_TX_DATE(最近交易日, 在考核区间内)', False),
    ('贷款维度(0062/0066)', 'DWD_ACCT_LOAN.CATE_5LVL: 1=正常/2=关注/3=次级/4=可疑/5=损失\n排除: 普惠贷款+微粒贷+美团联合贷款', False),
    ('所属机构(ORGN_ID)', 'DWD_CUST_MAN.ORGN_ID — 客户管户所属机构\nDWD_MKT_TSK_INFO.MKT_PERSN_ORG — 任务客户管户所属机构', False),
]

row = 5
for label, content, is_section in dim_rows:
    if is_section:
        section_row(ws5, row, 2, label, cat_fill)
    else:
        style_cell(ws5, row, 1, label, bold_font)
        style_cell(ws5, row, 2, content)
    row += 1

set_widths(ws5, [30, 90])

# Save
output = r'd:\AI\AI-Workspace\Kingbase-CRM-AI-Development-Guide\requirements\指标业务理解与技术口径_v3.xlsx'
wb.save(output)
print(f'Saved: {output}')