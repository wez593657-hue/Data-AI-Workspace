"""Generate ADS Kingbase DDL and data dictionaries from the ADS mapping workbook."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "data_assets" / "mapping" / "dws_to_ads" / "ADS应用层数据模型_CRM_ V1.0.xlsx"
DDL_DIR = ROOT / "data_assets" / "ddl" / "ads"
DICTIONARY_DIR = ROOT / "data_assets" / "data_dictionary" / "ads"
FIRST_ENTITY_SHEET = 3
LAST_ENTITY_SHEET = -1


@dataclass(frozen=True)
class Column:
    name: str
    source_type: str
    kingbase_type: str
    length: str
    comment: str
    primary_key: str
    enum: str


@dataclass(frozen=True)
class Entity:
    sheet_name: str
    table_name: str
    chinese_name: str
    columns: tuple[Column, ...]


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def sql_literal(value: str) -> str:
    return value.replace("'", "''")


def parse_type(raw_type: str, size: object, scale: object) -> tuple[str, str, str]:
    raw_type = raw_type.upper().replace(" ", "")
    if not raw_type:
        raise ValueError("缺少数据类型")

    matched = re.fullmatch(r"(VARCHAR2?|CHAR|NUMBER|DECIMAL|NUMERIC)\((\d+)(?:,(\d+))?\)", raw_type)
    if matched:
        family, precision, parsed_scale = matched.groups()
        if family in {"VARCHAR", "VARCHAR2"}:
            return "VARCHAR2", f"VARCHAR({precision})", precision
        if family == "CHAR":
            return "CHAR", f"CHAR({precision})", precision
        normalized = "NUMBER" if family == "NUMBER" else family
        precision_scale = f"{precision},{parsed_scale}" if parsed_scale else precision
        return normalized, f"{normalized}({precision_scale})", precision_scale

    if raw_type in {"VARCHAR", "VARCHAR2", "CHAR"}:
        length = text(size)
        if not length.isdigit():
            raise ValueError(f"{raw_type} 缺少有效长度: {size!r}")
        source_family = "VARCHAR2" if raw_type in {"VARCHAR", "VARCHAR2"} else "CHAR"
        kingbase_family = "VARCHAR" if source_family == "VARCHAR2" else "CHAR"
        return source_family, f"{kingbase_family}({length})", length

    if raw_type in {"NUMBER", "DECIMAL", "NUMERIC"}:
        precision = text(size)
        decimal_scale = text(scale)
        if not precision.isdigit():
            raise ValueError(f"{raw_type} 缺少有效精度: {size!r}")
        if decimal_scale and not decimal_scale.isdigit():
            raise ValueError(f"{raw_type} 精度位无效: {scale!r}")
        number_type = "NUMBER" if raw_type == "NUMBER" else raw_type
        precision_scale = f"{precision},{decimal_scale}" if decimal_scale else precision
        return number_type, f"{number_type}({precision_scale})", precision_scale

    if raw_type in {"DATE", "TIMESTAMP", "TEXT", "CLOB"}:
        return raw_type, raw_type, "-"
    raise ValueError(f"不支持的数据类型: {raw_type}")


def parse_entities() -> tuple[Entity, ...]:
    workbook = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=False)
    entities: list[Entity] = []
    tables: set[str] = set()
    for sheet_name in workbook.sheetnames[FIRST_ENTITY_SHEET:LAST_ENTITY_SHEET]:
        sheet = workbook[sheet_name]
        chinese_name = text(sheet.cell(2, 3).value)
        table_name = text(sheet.cell(2, 6).value).upper()
        if not chinese_name or not table_name:
            raise ValueError(f"{sheet_name}: 缺少实体中文名或物理表名")
        if table_name in tables:
            raise ValueError(f"物理表名重复: {table_name}")
        tables.add(table_name)

        columns: list[Column] = []
        names: set[str] = set()
        for row in range(6, sheet.max_row + 1):
            name = text(sheet.cell(row, 2).value).upper()
            if not name or name == "变更登记":
                break
            if name in names:
                raise ValueError(f"{sheet_name}: 字段重复: {name}")
            names.add(name)
            source_type, kingbase_type, length = parse_type(
                text(sheet.cell(row, 3).value) or text(sheet.cell(row, 4).value),
                sheet.cell(row, 4).value if text(sheet.cell(row, 3).value) else None,
                sheet.cell(row, 6).value,
            )
            columns.append(
                Column(
                    name=name,
                    source_type=source_type,
                    kingbase_type=kingbase_type,
                    length=length,
                    comment=text(sheet.cell(row, 5).value),
                    primary_key=text(sheet.cell(row, 8).value),
                    enum=text(sheet.cell(row, 10).value),
                )
            )
        if not columns:
            raise ValueError(f"{sheet_name}: 未解析到字段")
        entities.append(Entity(sheet_name, table_name, chinese_name, tuple(columns)))
    return tuple(entities)


def render_ddl(entity: Entity) -> str:
    lines = [
        "/*",
        f" * {entity.table_name}",
        f" * 中文名称: {entity.chinese_name}",
        " * 版本: v1.0",
        f" * 创建时间: {date.today().isoformat()}",
        " */",
        "",
        f"CREATE TABLE IF NOT EXISTS {entity.table_name} (",
    ]
    lines.extend(
        f"    {column.name} {column.kingbase_type}{',' if index < len(entity.columns) - 1 else ''}"
        for index, column in enumerate(entity.columns)
    )
    lines.extend([" );", "", f"COMMENT ON TABLE {entity.table_name} IS '{sql_literal(entity.chinese_name)}';"])
    lines.extend(
        f"COMMENT ON COLUMN {entity.table_name}.{column.name} IS '{sql_literal(column.comment)}';"
        for column in entity.columns
        if column.comment
    )
    return "\n".join(lines).replace("\n );", "\n);") + "\n"


def render_dictionary(entity: Entity) -> str:
    today = date.today().isoformat()
    lines = [
        f"# ADS数据字典 - {entity.table_name}",
        "",
        "## 表信息",
        "",
        "| 属性 | 值 |",
        "| --- | --- |",
        "| 层级 | ADS - 应用数据层 |",
        f"| 表名 | {entity.table_name} |",
        f"| 中文名称 | {entity.chinese_name} |",
        f"| 来源模型 | ADS应用层数据模型_CRM_ V1.0.xlsx / {entity.sheet_name} |",
        f"| 更新时间 | {today} |",
        "",
        "## 字段列表",
        "",
        "| 字段名 | 字段中文说明 | 数据类型 | 长度 | 是否为空 | 默认值 | 主键 | 外键 | 枚举说明 | 业务含义 |",
        "| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for column in entity.columns:
        primary_key = column.primary_key if column.primary_key else "-"
        enum = column.enum.replace("|", "\\|") if column.enum else "-"
        comment = column.comment.replace("|", "\\|") if column.comment else "-"
        lines.append(
            f"| {column.name} | {comment} | {column.source_type} | {column.length} | 【待确认】 | - | {primary_key} | - | {enum} | {comment} |"
        )
    lines.extend(["", "---", "", f"*数据字典版本: v1.0 | 生成时间: {today}*"])
    return "\n".join(lines) + "\n"


def validate_outputs(entities: tuple[Entity, ...]) -> None:
    expected_names = {entity.table_name.lower() for entity in entities}
    actual_ddl = {path.stem for path in DDL_DIR.glob("*.sql")}
    actual_dictionary = {path.stem for path in DICTIONARY_DIR.glob("*.md")}
    if actual_ddl != expected_names:
        raise ValueError(f"DDL 文件集合不一致: {sorted(actual_ddl ^ expected_names)}")
    if actual_dictionary != expected_names:
        raise ValueError(f"数据字典文件集合不一致: {sorted(actual_dictionary ^ expected_names)}")
    for entity in entities:
        filename = entity.table_name.lower()
        ddl_path = DDL_DIR / f"{filename}.sql"
        dictionary_path = DICTIONARY_DIR / f"{filename}.md"
        if ddl_path.read_text(encoding="utf-8") != render_ddl(entity):
            raise ValueError(f"{entity.table_name}: DDL 内容与映射文件不一致")
        if dictionary_path.read_text(encoding="utf-8") != render_dictionary(entity):
            raise ValueError(f"{entity.table_name}: 数据字典内容与映射文件不一致")


def main() -> None:
    entities = parse_entities()
    DDL_DIR.mkdir(parents=True, exist_ok=True)
    DICTIONARY_DIR.mkdir(parents=True, exist_ok=True)
    expected_names = {entity.table_name.lower() for entity in entities}
    for directory, suffix in ((DDL_DIR, ".sql"), (DICTIONARY_DIR, ".md")):
        for path in directory.glob(f"ads_*{suffix}"):
            if path.stem not in expected_names:
                path.unlink()

    for entity in entities:
        filename = entity.table_name.lower()
        (DDL_DIR / f"{filename}.sql").write_text(render_ddl(entity), encoding="utf-8")
        (DICTIONARY_DIR / f"{filename}.md").write_text(render_dictionary(entity), encoding="utf-8")
    validate_outputs(entities)
    print(f"已生成并验证 {len(entities)} 张 ADS 表的 DDL 和数据字典。")


if __name__ == "__main__":
    main()
