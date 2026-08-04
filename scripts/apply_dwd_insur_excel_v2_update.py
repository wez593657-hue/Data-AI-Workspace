# -*- coding: utf-8 -*-
"""
DWD_ACCT_INSUR v2.0 Excel 权威源更新脚本
==========================================
sheet 列结构（保险账户信息）：
  B=属性名称  C=属性类型  D=数据类型  E=标准中文名  F=属性校验规则
  L=源系统表名  M=源表别名  N=源系统字段英文名  O=源系统字段中文名
  P=映射规则  Q=备注

变更内容：
1. TX_DATE：类型 VARCHAR2(8)，中文名"首次交易日期"
2. 新增 LAST_TX_DATE / ACTL_TERM_DATE / NEW_INSUR_AMT
3. INSUR_AMT：= 新单保费+续期保费累计；终止/未生效/趸交满一年/期缴缴满/宽限期过60天未缴 均置0
4. POLICY_STATE：VARCHAR2(8) NOT NULL，值域 0未生效/1正常/2失效
5. CANCL_INSUR_DATE：保险期间结束日期(推算值,仅参考)
6. TX_TYP：置空（全链路不消费）
7. 变更登记追加 2026-08-03 v2.0 记录

前置条件：关闭 Excel。
幂等性：按字段名定位；已存在字段跳过新增，仅更新规则。
"""

import os
import sys

import openpyxl

# 列映射（1-based）
COL_NAME = 2
COL_ATTR_TYPE = 3
COL_DATA_TYPE = 4
COL_CN_NAME = 5
COL_RULE = 6
COL_SRC_TABLE = 12
COL_SRC_ALIAS = 13
COL_SRC_FIELD = 14
COL_SRC_FIELD_CN = 15
COL_MAP_RULE = 16
COL_REMARK = 17


def _blocked_cells(ws):
    blocked = set()
    for mr in ws.merged_cells.ranges:
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                if not (row == mr.min_row and col == mr.min_col):
                    blocked.add((row, col))
    return blocked


def _set(ws, blocked, row, col, value):
    if (row, col) in blocked or value is None:
        return
    ws.cell(row=row, column=col).value = value


def main():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    map_dir = os.path.join(base_dir, "data_assets", "mapping", "ods_to_dwd")
    xlsx = None
    for fn in os.listdir(map_dir):
        if fn.endswith(".xlsx") and not fn.startswith("~$"):
            xlsx = os.path.join(map_dir, fn)
            break
    if xlsx is None:
        print("ERR: DWD 数据模型 Excel 未找到")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx, data_only=False)
    ws = wb["\u4fdd\u9669\u8d26\u6237\u4fe1\u606f"]  # 保险账户信息
    blocked = _blocked_cells(ws)

    rows = {}
    for r in range(1, ws.max_row + 1):
        name = ws.cell(row=r, column=COL_NAME).value
        if name:
            rows[str(name).strip()] = r

    # 1) TX_DATE
    if "TX_DATE" in rows:
        r = rows["TX_DATE"]
        _set(ws, blocked, r, COL_DATA_TYPE, "VARCHAR2(8)")
        _set(ws, blocked, r, COL_CN_NAME, "\u9996\u6b21\u4ea4\u6613\u65e5\u671f")
        _set(ws, blocked, r, COL_REMARK,
             "\u9996\u6b21\u7f34\u8d39\u65e5\uff0c\u8db4\u4ea4\u6ee1\u4e00\u5e74\u8d77\u7b97\u57fa\u51c6(YYYYMMDD)")

    # 2) CANCL_INSUR_DATE
    if "CANCL_INSUR_DATE" in rows:
        r = rows["CANCL_INSUR_DATE"]
        _set(ws, blocked, r, COL_CN_NAME, "\u4fdd\u9669\u671f\u95f4\u7ed3\u675f\u65e5\u671f(\u63a8\u7b97\u503c,\u4ec5\u53c2\u8003)")
        _set(ws, blocked, r, COL_REMARK,
             "\u4e0d\u53c2\u4e0e\u7ec8\u6b62\u5224\u5b9a\uff1b\u7ec8\u6b62\u5224\u5b9a\u7528 ACTL_TERM_DATE")

    # 3) POLICY_STATE
    if "POLICY_STATE" in rows:
        r = rows["POLICY_STATE"]
        _set(ws, blocked, r, COL_DATA_TYPE, "VARCHAR2(8)")
        _set(ws, blocked, r, COL_RULE,
             "0\u672a\u751f\u6548/1\u6b63\u5e38/2\u5931\u6548\uff1bNOT NULL\uff0c\u975e0/1/2\u503c\u62e6\u622a")
        _set(ws, blocked, r, COL_REMARK,
             "\u552f\u4e00\u72b6\u6001\u5224\u5b9a\u6e90\uff1bDWS/ADS \u4e0d\u518d\u4f9d\u636e\u4ea4\u6613\u7c7b\u578b")

    # 4) INSUR_AMT
    if "INSUR_AMT" in rows:
        r = rows["INSUR_AMT"]
        _set(ws, blocked, r, COL_CN_NAME, "\u5f53\u524d\u4fdd\u9669\u91d1\u989d")
        _set(ws, blocked, r, COL_RULE, "INSUR_AMT >= 0")
        _set(ws, blocked, r, COL_MAP_RULE,
             "NEW_INSUR_AMT + \u7eed\u671f\u4fdd\u8d39\u7d2f\u8ba1\uff1b\u7ec8\u6b62/\u672a\u751f\u6548/\u8db4\u4ea4\u6ee1\u4e00\u5e74/\u671f\u7f34\u7f34\u6ee1/\u5bbd\u9650\u671f\u8fc760\u5929\u672a\u7f34 \u5747\u7f6e0")
        _set(ws, blocked, r, COL_REMARK,
             "\u8db4\u4ea4\u6ee1\u4e00\u5e74\u8d77\u7b97\u70b9=TX_DATE\uff1b60\u5929\u5bbd\u9650\u671f\u89c4\u5219\u7eb3\u5165")

    # 5) TX_TYP
    if "TX_TYP" in rows:
        r = rows["TX_TYP"]
        _set(ws, blocked, r, COL_MAP_RULE, "\u7f6e\u7a7a\uff08v2.0 \u8d77\u5168\u94fe\u8def\u4e0d\u6d88\u8d39\uff09")
        _set(ws, blocked, r, COL_REMARK,
             "\u72b6\u6001\u5224\u5b9a\u4ec5\u7528 POLICY_STATE/ACTL_TERM_DATE")

    # 6) 新增 3 字段（PERSN_LEGAL_BK_CODE 之后第一个空行起）
    new_fields = [
        ["LAST_TX_DATE", "VARCHAR2(8)", "\u6700\u8fd1\u4ea4\u6613\u65e5\u671f", "LAST_TX_DATE >= TX_DATE",
         "YBT_POLICY_FEE_LIST", "a", "a.TX_DATE", "\u4ea4\u6613\u65e5\u671f",
         "MAX(\u4ea4\u6613\u65e5\u671f)",
         "\u4fdd\u5355\u6700\u8fd1\u4e00\u7b14\u4ea4\u6613\u65e5\u671f(YYYYMMDD)\uff1bODS\u5b57\u6bb5\u540d\u5f85\u6838\u5bf9"],
        ["ACTL_TERM_DATE", "VARCHAR2(8)", "\u5b9e\u9645\u7ec8\u6b62\u65e5\u671f", "ACTL_TERM_DATE >= TX_DATE",
         "YBT_POLICY_FEE_LIST", "a", "\u72b6\u6001\u4ea4\u6613\u65e5\u671f", "\u7ec8\u6b62\u4ea4\u6613\u65e5\u671f",
         "\u72b6\u6001\u4ea4\u6613(2/3/4/5/6/8)\u6700\u65b0\u65e5\u671f\u2192CONT_STATUS\u7ec8\u6b62\u53d8\u66f4\u65e5\u2192CANCL_INSUR_DATE\u56de\u9000",
         "\u7ec8\u6b62\u5224\u5b9a\u4e0e\u7edf\u8ba1\uff1b\u7ec8\u6b62\u65f6\u975e\u7a7a"],
        ["NEW_INSUR_AMT", "NUMBER(20,2)", "\u9996\u671f\u4fdd\u8d39(\u65b0\u5355\u4fdd\u8d39)", "NEW_INSUR_AMT >= 0",
         "YBT_POLICY_FEE_LIST", "a", "a.ORD_AMT", "\u8ba2\u5355\u603b\u4fdd\u8d39",
         "TRAN_TYPE='0' \u65f6\u7684 ORD_AMT",
         "\u652f\u6491\u6307\u68070061\u4e0e ADS.INSUR_FRST_PREM_AMT"],
    ]
    anchor = rows.get("PERSN_LEGAL_BK_CODE")
    r = anchor + 1 if anchor else ws.max_row + 1
    while ws.cell(row=r, column=COL_NAME).value is not None:
        r += 1
    cols = [COL_NAME, COL_DATA_TYPE, COL_CN_NAME, COL_RULE,
            COL_SRC_TABLE, COL_SRC_ALIAS, COL_SRC_FIELD, COL_SRC_FIELD_CN,
            COL_MAP_RULE, COL_REMARK]
    for vals in new_fields:
        if vals[0] in rows:
            continue
        for c, v in zip(cols, vals):
            _set(ws, blocked, r, c, v)
        r += 1

    # 7) 变更登记（B=日期 C=修改人 D=备注）
    change_title = None
    for rr in range(1, ws.max_row + 1):
        if str(ws.cell(row=rr, column=2).value or "").strip() == "\u53d8\u66f4\u767b\u8bb0":
            change_title = rr
            break
    if change_title:
        rr = change_title + 1
        # 幂等：若已存在 2026-08-03 v2.0 登记则跳过
        exist = False
        while rr <= ws.max_row:
            b = ws.cell(row=rr, column=2).value
            if b and "2026-08-03" in str(b):
                exist = True
                break
            if b is None and ws.cell(row=rr, column=3).value is None:
                break
            rr += 1
        if not exist:
            rr = change_title + 1
            while ws.cell(row=rr, column=2).value is not None:
                rr += 1
            _set(ws, blocked, rr, 2, "2026-08-03")
            _set(ws, blocked, rr, 3, "-")
            _set(ws, blocked, rr, 4,
                 "v2.0 \u4fdd\u5355\u7ea7\u4e3b\u6863\uff1a\u65b0\u589e LAST_TX_DATE/ACTL_TERM_DATE/NEW_INSUR_AMT\uff1bTX_DATE=\u9996\u6b21\u4ea4\u6613\u65e5\uff1bPOLICY_STATE(0/1/2)\u72b6\u6001\u5224\u5b9a\uff1bINSUR_AMT\u7ec8\u6b62/\u7f34\u8d39\u671f\u6ee1/\u5bbd\u9650\u671f\u8fc7\u6e05\u96f6")

    wb.save(xlsx)
    print("SAVED:", xlsx)


if __name__ == "__main__":
    main()
