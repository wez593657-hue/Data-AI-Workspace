# -*- coding: utf-8 -*-
"""更新 xlsx: 0068 改从标签表 ADS_CUST_LABEL_BILL_INFO 取 + sheet4 加表行"""
import openpyxl
from openpyxl.styles import Alignment

XLSX = r'requirements/指标业务理解与技术口径.xlsx'
wb = openpyxl.load_workbook(XLSX)
WRAP = Alignment(wrap_text=True, vertical='top')

def setcell(ws, row, col, val):
    ws.cell(row=row, column=col).value = val
    ws.cell(row=row, column=col).alignment = WRAP

ws = wb['指标技术规格']
# R27 = 0068 行
setcell(ws, 27, 8,  "CURNT_VAL = COUNT(DISTINCT CUST_ID)\n取标签表 ADS_CUST_LABEL_BILL_INFO.IS_NOT_BILL_RSV_VAL_MKNT='1'（收单价值商户, 当天快照）\n【2026-08-18 更新】不再自算资产+交易双门槛，由上游标签加工")
setcell(ws, 27, 9,  "ADS_CUST_LABEL_BILL_INFO (收单商户标签, 当天快照)\n字段: IS_NOT_BILL_RSV_VAL_MKNT(价值商户) / IS_NOT_BILL_RSV_MINOR_MKNT(小微) / IS_NOT_BILL_RSV_INDIV_MKNT(个体)")
setcell(ws, 27, 12, "【口径2026-08-18】收单价值商户 = 标签表 IS_NOT_BILL_RSV_VAL_MKNT='1'\n[源表] ADS_CUST_LABEL_BILL_INFO\n[DDL] 待补建 data_assets/ddl/ads/ads_cust_label_bill_info.sql")
setcell(ws, 27, 13, "【待确认】0069 人群是否同取该表(小微/个体字段)【待确认】；标签表 DDL 需补建")
setcell(ws, 27, 14, "❌ 未开发 (口径已定, 待实现)")
setcell(ws, 27, 15, "—（未实现）")
setcell(ws, 27, 16, "—\n依赖 ADS_CUST_LABEL_BILL_INFO(DDL待补建)，已定方案见规划")

# sheet4 数据源映射追加
ws4 = wb['数据源映射']
newrow = 31
for c, v in enumerate([
    "ADS_CUST_LABEL_BILL_INFO", "ADS", "0068",
    "PERSN_LEGAL_BK_CODE,DATA_DATE,CUST_ID,IS_NOT_BILL_RSV_VAL_MKNT(价值),IS_NOT_BILL_RSV_MINOR_MKNT(小微),IS_NOT_BILL_RSV_INDIV_MKNT(个体)",
    "收单商户标签: IS_NOT_BILL_RSV_VAL_MKNT='1' 为收单价值商户(0068数据源)",
    "待补建", "2026-08-18 用户确认 0068 改从此表取；DDL 待建"
], start=1):
    setcell(ws4, newrow, c, v)

wb.save(XLSX)
print("XLSX SAVED OK")