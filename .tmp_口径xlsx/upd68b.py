# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Alignment
XLSX = r'.tmp_口径xlsx/work_v42.xlsx'
wb = openpyxl.load_workbook(XLSX)
WRAP = Alignment(wrap_text=True, vertical='top')
def sc(ws, r, c, v):
    ws.cell(row=r, column=c).value = v
    ws.cell(row=r, column=c).alignment = WRAP

H_8 = ("CURNT_VAL = COUNT(DISTINCT CUST_ID)\n"
       "取标签表 ADS_CUST_LABEL_BILL_INFO.IS_NOT_BILL_RSV_VAL_MKNT='1'（收单价值商户, 当天快照）\n"
       "【2026-08-18 更新】不再自算资产+交易双门槛")
I_9 = ("ADS_CUST_LABEL_BILL_INFO (收单商户标签, 当天快照)\n"
       "字段: IS_NOT_BILL_RSV_VAL_MKNT(价值)/IS_NOT_BILL_RSV_MINOR_MKNT(小微)/IS_NOT_BILL_RSV_INDIV_MKNT(个体)")
L_12 = ("【口径2026-08-18】价值商户 = 标签表 IS_NOT_BILL_RSV_VAL_MKNT='1'\n"
        "[源表] ADS_CUST_LABEL_BILL_INFO\n[DDL] 待补建 data_assets/ddl/ads/ads_cust_label_bill_info.sql")
M_13 = "【待确认】0069 人群是否同取该表(小微/个体字段)；标签表 DDL 需补建"
ws = wb['指标技术规格']
sc(ws, 27, 8, H_8)
sc(ws, 27, 9, I_9)
sc(ws, 27, 12, L_12)
sc(ws, 27, 13, M_13)
sc(ws, 27, 14, "❌ 未开发 (口径已定, 待实现)")
sc(ws, 27, 15, "—（未实现）")
sc(ws, 27, 16, "—\n依赖 ADS_CUST_LABEL_BILL_INFO(DDL待补建)")

ws4 = wb['数据源映射']
row = ["ADS_CUST_LABEL_BILL_INFO", "ADS", "0068",
       "PERSN_LEGAL_BK_CODE,DATA_DATE,CUST_ID,IS_NOT_BILL_RSV_VAL_MKNT(价值),IS_NOT_BILL_RSV_MINOR_MKNT(小微),IS_NOT_BILL_RSV_INDIV_MKNT(个体)",
       "收单商户标签: IS_NOT_BILL_RSV_VAL_MKNT='1'为收单价值商户(0068数据源)",
       "待补建", "2026-08-18 0068改从此表取; DDL待建"]
for c, v in enumerate(row, start=1):
    sc(ws4, 31, c, v)
wb.save(XLSX)
print("TMP SAVED OK")
